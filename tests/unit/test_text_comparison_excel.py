from __future__ import annotations

import io

from openpyxl import load_workbook

from vigilance.text_comparison.justification import build_text_triage_justification
from vigilance.text_comparison.text_comparison_excel import generate_text_comparison_excel


def _column(ws, header: str) -> int:
    """Retourner l'index d'une colonne analyste par son en-tête."""
    for index in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=index).value == header:
            return index
    raise AssertionError(f"Colonne absente : {header}")


def test_generate_text_comparison_excel_creates_analysis_sheet() -> None:
    justification_oui = (
        "OUI - le nouveau modele AIRB est introduit au T2 et n'apparaissait "
        "pas au T1. Cette methode change la facon dont la banque decrit "
        "l'evaluation du risque de credit.\n\n"
        "Cette nouveaute est pertinente pour la vigie parce qu'elle touche "
        "une methodologie prudentielle et les exigences BSIF, pas seulement "
        "une reformulation.\n\n"
        "L'analyste doit comparer cette nouvelle base methodologique avec "
        "celle du trimestre precedent et verifier l'impact sur la comparabilite "
        "inter-pairs."
    )
    justification_non = (
        "NON - la valeur du ratio change mais l'indicateur existait deja au "
        "T1. Le T2 ne cree pas de nouveau concept ni de nouvelle methode.\n\n"
        "Cette variation chiffree est propre a la banque et ne touche pas un "
        "seuil reglementaire AMF ou BSIF. Elle ne constitue donc pas une "
        "nouvelle idee de vigie.\n\n"
        "L'analyste peut l'ecarter comme mise a jour quantitative attendue, "
        "sauf si un seuil interne distinct declenche une revue separee."
    )

    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "block_comparisons": [],
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "semantic_text_t1": "Ancien texte",
                        "semantic_text_t2": "Nouveau texte majeur",
                        "source_text_t1": "Paragraphe exact T1",
                        "source_text_t2": "Paragraphe exact T2",
                        "evidence_t1": {"pages": [10], "snippet": "preuve T1"},
                        "evidence_t2": {"pages": [12], "snippet": "preuve"},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "RISQUE",
                            "impact_level": "MAJEUR",
                            "action_requise": "revue_prioritaire",
                            "nouvelle_idee": True,
                            "nouvelle_idee_justification": justification_oui,
                        },
                    },
                    {
                        "diff_type": "added",
                        "semantic_text_t1": "",
                        "semantic_text_t2": "Texte modere expert",
                        "source_text_t1": "",
                        "source_text_t2": "Paragraphe exact ajouté",
                        "evidence_t2": {"pages": [13], "snippet": "preuve moderee"},
                        "genai_triage": {
                            "is_relevant": False,
                            "category": "NON_PERTINENT",
                            "impact_level": "MINEUR",
                            "action_requise": "aucune",
                            "nouvelle_idee": False,
                            "nouvelle_idee_justification": "",
                        },
                    },
                    {
                        "diff_type": "modified",
                        "semantic_text_t1": "Texte modéré",
                        "semantic_text_t2": "Texte modéré substantif",
                        "source_text_t1": "Paragraphe modéré T1",
                        "source_text_t2": "Paragraphe modéré T2",
                        "evidence_t1": {"pages": [14], "snippet": "preuve moderee T1"},
                        "evidence_t2": {"pages": [15], "snippet": "preuve moderee T2"},
                        "genai_triage": {
                            "is_relevant": False,
                            "category": "NON_PERTINENT",
                            "impact_level": "MODERE",
                            "action_requise": "information",
                            "nouvelle_idee": False,
                            "nouvelle_idee_justification": "",
                        },
                    },
                    {
                        "diff_type": "modified",
                        "semantic_text_t1": "Texte non substantif T1",
                        "semantic_text_t2": "Texte non substantif T2",
                        "source_text_t1": "Paragraphe non substantif T1",
                        "source_text_t2": "Paragraphe non substantif T2",
                        "evidence_t1": {"pages": [16], "snippet": "preuve non substantif T1"},
                        "evidence_t2": {"pages": [17], "snippet": "preuve non substantif T2"},
                        "genai_triage": {
                            "is_relevant": False,
                            "category": "NON_PERTINENT",
                            "impact_level": "MINEUR",
                            "action_requise": "aucune",
                            "nouvelle_idee": False,
                            "nouvelle_idee_justification": "",
                        },
                    },
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))

    assert workbook.sheetnames == ["Analyse complète"]
    ws = workbook["Analyse complète"]
    assert ws.max_row == 8
    rows = [
        {
            "current": ws.cell(row, _column(ws, "Texte exact du trimestre courant")).value,
            "previous": ws.cell(row, _column(ws, "Texte exact du trimestre précédent")).value,
            "type": ws.cell(row, _column(ws, "Type de changement")).value,
            "section": ws.cell(row, _column(ws, "Section du rapport")).value,
            "current_page": ws.cell(row, _column(ws, "Page du texte courant")).value,
            "previous_page": ws.cell(row, _column(ws, "Page du texte précédent")).value,
        }
        for row in range(2, ws.max_row + 1)
    ]
    assert {row["type"] for row in rows} <= {"Ajout", "Suppression", "Renommage"}
    replacement_rows = [row for row in rows if row["current"] == "Paragraphe exact T2"]
    assert {row["type"] for row in replacement_rows} == {"Ajout", "Suppression"}
    assert all(row["previous"] == "Paragraphe exact T1" for row in replacement_rows)
    assert all(row["section"] == "Gestion des risques" for row in replacement_rows)
    assert all(row["current_page"] == "12" and row["previous_page"] == "10" for row in replacement_rows)
    assert any(
        row["type"] == "Ajout"
        and row["current"] == "Paragraphe exact ajouté"
        and row["previous"] is None
        for row in rows
    )


def test_generate_text_comparison_excel_keeps_minor_date_and_reformulation_changes() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "change_summary": "La date de référence est passée de janvier à avril.",
                        "source_text_t1": "Données au 31 janvier.",
                        "source_text_t2": "Données au 30 avril.",
                        "evidence_t1": {"pages": [4]},
                        "evidence_t2": {"pages": [5]},
                        "genai_triage": {
                            "is_relevant": False,
                            "category": "NON_PERTINENT",
                            "impact_level": "MINEUR",
                            "action_requise": "aucune",
                            "nouvelle_idee": False,
                            "exclusion_reason": "variation_numerique_propre_banque",
                            "nouvelle_idee_justification": (
                                "NON - changement de date seulement."
                            ),
                        },
                    },
                    {
                        "diff_type": "modified",
                        "change_summary": "Légère reformulation sans changement de fond.",
                        "source_text_t1": "La banque surveille ce risque.",
                        "source_text_t2": "Ce risque est surveillé par la banque.",
                        "evidence_t1": {"pages": [6]},
                        "evidence_t2": {"pages": [7]},
                        "genai_triage": {
                            "is_relevant": False,
                            "category": "NON_PERTINENT",
                            "impact_level": "MINEUR",
                            "action_requise": "aucune",
                            "nouvelle_idee": False,
                            "exclusion_reason": "reformulation_mineure",
                            "nouvelle_idee_justification": (
                                "NON - reformulation sans changement de sens."
                            ),
                        },
                    },
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]

    assert ws.max_row == 5
    previous_values = [
        ws.cell(row, _column(ws, "Texte exact du trimestre précédent")).value
        for row in range(2, ws.max_row + 1)
    ]
    current_values = [
        ws.cell(row, _column(ws, "Texte exact du trimestre courant")).value
        for row in range(2, ws.max_row + 1)
    ]
    assert previous_values.count("Données au 31 janvier.") == 2
    assert current_values.count("Données au 30 avril.") == 2
    assert previous_values.count("La banque surveille ce risque.") == 2
    assert current_values.count("Ce risque est surveillé par la banque.") == 2


def test_generate_text_comparison_excel_strips_control_characters() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "source_text_t1": "Texte\x00 T1",
                        "source_text_t2": "Texte T2",
                        "evidence_t1": {"pages": [1]},
                        "evidence_t2": {"pages": [2]},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "RISQUE",
                            "impact_level": "MAJEUR",
                            "action_requise": "revue_prioritaire",
                            "nouvelle_idee": True,
                            "nouvelle_idee_justification": "OUI - note\x00 analyste utile.",
                        },
                    }
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]

    assert ws.cell(2, _column(ws, "Texte exact du trimestre précédent")).value == "Texte T1"
    assert "\x00" not in ws.cell(2, _column(ws, "Justification de pertinence (IA)")).value
    assert "note analyste utile" in ws.cell(2, _column(ws, "Justification de pertinence (IA)")).value


def test_generate_text_comparison_excel_applies_analyst_review_without_new_columns() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "change_id": "chg-rejected",
                        "diff_type": "modified",
                        "source_text_t1": "Ancien",
                        "source_text_t2": "Nouveau",
                        "evidence_t1": {"pages": [1]},
                        "evidence_t2": {"pages": [2]},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "RISQUE",
                            "impact_level": "MAJEUR",
                            "nouvelle_idee": True,
                            "nouvelle_idee_justification": "OUI - idée nouvelle.",
                        },
                        "_analyst_review": {
                            "status": "rejected",
                            "comment": "Pas une nouvelle idée.",
                        },
                    },
                    {
                        "change_id": "chg-approved",
                        "diff_type": "added",
                        "source_text_t1": "",
                        "source_text_t2": "Ajout",
                        "evidence_t2": {"pages": [3]},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "RISQUE",
                            "impact_level": "MAJEUR",
                            "nouvelle_idee": True,
                            "nouvelle_idee_justification": "OUI - ajout.",
                        },
                        "_analyst_review": {
                            "status": "approved",
                            "comment": "À conserver.",
                        },
                    },
                    {
                        "change_id": "chg-skipped",
                        "diff_type": "modified",
                        "source_text_t1": "A",
                        "source_text_t2": "B",
                        "evidence_t1": {"pages": [4]},
                        "evidence_t2": {"pages": [5]},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "RISQUE",
                            "impact_level": "MAJEUR",
                            "nouvelle_idee": True,
                            "nouvelle_idee_justification": "OUI - autre.",
                        },
                        "_analyst_review": {
                            "status": "skipped",
                            "comment": "Ne doit pas sortir.",
                        },
                    },
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    assert headers == [
        "Texte exact du trimestre courant",
        "Texte exact du trimestre précédent",
        "Catégorie principale",
        "Étiquettes secondaires",
        "Section du rapport",
        "Sous-section",
        "Type d'élément",
        "Type de changement",
        "Ce qui change",
        "Nouvelle idée à surveiller ?",
        "Justification de pertinence (IA)",
        "Priorité / impact",
        "Page du texte courant",
        "Page du texte précédent",
        "Statut analyste",
        "Note analyste",
        "Validé le",
    ]
    values = {
        ws.cell(row=row, column=_column(ws, "Texte exact du trimestre courant")).value: (
            ws.cell(row=row, column=_column(ws, "Nouvelle idée à surveiller ?")).value,
            ws.cell(row=row, column=_column(ws, "Note analyste")).value,
            ws.cell(row=row, column=_column(ws, "Statut analyste")).value,
        )
        for row in range(2, ws.max_row + 1)
    }
    assert values["Nouveau"] == ("Non", "Pas une nouvelle idée.", "Rejeté")
    assert values["Ajout"] == ("Oui", "À conserver.", "Validé")
    assert values["B"] == ("Oui", None, "Ignoré")


def test_generate_text_comparison_excel_excludes_confirmed_moves() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "alignment_decision": "moved_text",
                        "source_text_t1": "Texte déplacé avant.",
                        "source_text_t2": "Texte déplacé après.",
                        "genai_triage": {"nouvelle_idee": False},
                    }
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]

    assert ws.max_row == 1


def test_text_justification_falls_back_for_legacy_b15_triage() -> None:
    change = {
        "diff_type": "added",
        "change_summary": (
            "Sous-section ajoutee: Faits nouveaux sur la reglementation en "
            "matiere de durabilite"
        ),
        "source_text_t1": "",
        "source_text_t2": (
            "En mars 2025, le BSIF a publie une mise a jour de la ligne "
            "directrice B-15 afin de l'harmoniser avec les exigences des "
            "normes du Conseil canadien des normes d'information sur la "
            "durabilite. Les principales modifications comprennent le report "
            "de la date de mise en oeuvre pour la communication d'informations "
            "sur les emissions de GES du champ d'application 3."
        ),
        "evidence_t2": {"pages": [43]},
        "genai_triage": {
            "is_relevant": True,
            "category": "REGLEMENTAIRE",
            "impact_level": "MODERE",
            "risk_type": "conformite",
            "action_requise": "confirmation",
            "nouvelle_idee": True,
            "explanation": (
                "Une nouvelle section a ete ajoutee concernant la mise a jour "
                "des lignes directrices pour s'aligner sur les normes de "
                "durabilite."
            ),
            "impact_description": (
                "Le report des exigences de communication sur les emissions "
                "de GES permet a la banque de mieux se preparer"
            ),
        },
    }

    justification = build_text_triage_justification(change)

    assert justification.startswith("OUI — Nouvel élément à surveiller : Oui.")
    assert "Sujet détecté : Risque climatique, ESG" in justification
    assert "nouvelle mention réglementaire" in justification
    assert "information ajoutée" in justification
    assert "Ce qui change : Le rapport courant ajoute" in justification
    assert "Pertinence métier :" in justification
    assert (
        "Ce changement met l'accent sur l'évolution du cadre réglementaire "
        "applicable aux divulgations climatiques des institutions financières."
    ) in justification
    assert (
        "Ce point est important à suivre, car il permet d'évaluer l'évolution "
        "des attentes prudentielles, la comparabilité des pratiques de divulgation "
        "entre pairs et le niveau de préparation des banques face aux exigences "
        "climatiques."
    ) in justification
    assert "Point de surveillance :" in justification

    raw = generate_text_comparison_excel(
        {
            "section_comparisons": [
                {
                    "section_key": "gestion_reglementation",
                    "section_title": "Faits nouveaux en matiere de reglementation",
                    "all_block_comparisons": [change],
                }
            ]
        },
        output_path=None,
    )
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]
    assert ws.cell(2, _column(ws, "Nouvelle idée à surveiller ?")).value == "Oui"
    assert "mise a jour des lignes directrices" in ws.cell(2, _column(ws, "Justification de pertinence (IA)")).value


def test_text_justification_rewrites_b15_pertinence_when_explicit_exists() -> None:
    change = {
        "diff_type": "added",
        "source_text_t2": "Le BSIF met à jour la ligne directrice B-15 sur les GES.",
        "genai_triage": {
            "nouvelle_idee": True,
            "nouvelle_idee_justification": (
                "OUI — Nouvel élément à surveiller : Oui.\n\n"
                "Sujet détecté : Risque climatique, ESG.\n\n"
                "Ce qui change : Le T2 ajoute une mention B-15.\n\n"
                "Pertinence métier : Ancienne formulation à remplacer.\n\n"
                "Point de surveillance : Le point à retenir demeure le suivi des exigences."
            ),
        },
    }

    justification = build_text_triage_justification(change)

    assert "Ancienne formulation à remplacer" not in justification
    assert (
        "Pertinence métier : Ce changement met l'accent sur l'évolution du cadre "
        "réglementaire applicable aux divulgations climatiques des institutions "
        "financières."
    ) in justification
    assert (
        "Point de surveillance : Risque climatique / ESG — Le changement indique "
        "que la banque tient compte de la mise à jour de la ligne directrice B-15"
    ) in justification


def test_generate_text_comparison_excel_labels_text_renames() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "renamed",
                        "subsection_heading": "Ancien titre → Nouveau titre",
                        "source_text_t1": "Ancien titre",
                        "source_text_t2": "Nouveau titre",
                        "evidence_t1": {"pages": [4]},
                        "evidence_t2": {"pages": [5]},
                        "genai_triage": {
                            "is_relevant": True,
                            "category": "STRUCTURE",
                            "impact_level": "MODERE",
                            "nouvelle_idee": False,
                            "nouvelle_idee_justification": "Renommage de sous-section.",
                        },
                    }
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]

    assert ws.cell(2, _column(ws, "Sous-section")).value == "Ancien titre → Nouveau titre"
    assert ws.cell(2, _column(ws, "Type de changement")).value == "Renommage"
    assert ws.cell(2, _column(ws, "Texte exact du trimestre précédent")).value == "Ancien titre"
    assert ws.cell(2, _column(ws, "Texte exact du trimestre courant")).value == "Nouveau titre"


def test_generate_text_comparison_excel_uses_french_analyst_labels() -> None:
    payload = {
        "section_comparisons": [
            {
                "section_key": "gestion_capital",
                "section_title": "Gestion du capital",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "change_summary": (
                            "Les deux fragments traitent de la même divulgation "
                            "concernant le report du plancher par le BSIF."
                        ),
                        "source_text_t1": "Calendrier jusqu'en 2027.",
                        "source_text_t2": "Report jusqu'à nouvel ordre.",
                        "evidence_t1": {"pages": [10]},
                        "evidence_t2": {"pages": [11]},
                        "genai_triage": {
                            "is_relevant": True,
                            "themes_amf": ["FONDS_PROPRES_REGLEMENTAIRES"],
                            "impact_level": "MAJEUR",
                            "nouvelle_idee": True,
                            "relevance_reason": (
                                "Le rapport courant actualise le calendrier d'application "
                                "du plancher de fonds propres. Cette mise à jour n'apporte "
                                "pas de méthode nouvelle à comparer entre les banques."
                            ),
                        },
                    }
                ],
            }
        ]
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]

    impact_col = _column(ws, "Priorité / impact")
    what_col = _column(ws, "Ce qui change")
    just_col = _column(ws, "Justification de pertinence (IA)")

    for row in range(2, ws.max_row + 1):
        impact = str(ws.cell(row, impact_col).value or "")
        what = str(ws.cell(row, what_col).value or "")
        justification = str(ws.cell(row, just_col).value or "")
        assert impact in {"Majeur", "Modéré", "Mineur"}
        for forbidden in ("fragment", "chunk", "T1", "T2", "MAJEUR", "MODERE"):
            assert forbidden not in what
            assert forbidden not in justification
        assert "calendrier" in what.lower()
        assert "Les deux passages" not in what
        assert "Les deux fragments" not in what


def test_what_changed_for_display_prefers_relevance_reason() -> None:
    from vigilance.vigie_columns import what_changed_for_display

    change = {
        "diff_type": "modified",
        "change_summary": (
            "Les deux fragments traitent de la même divulgation "
            "concernant le report du plancher par le BSIF."
        ),
        "source_text_t1": "Calendrier jusqu'en 2027.",
        "source_text_t2": "Report jusqu'à nouvel ordre.",
        "genai_triage": {
            "relevance_reason": (
                "Le rapport courant actualise le calendrier d'application "
                "du plancher de fonds propres. Cette mise à jour n'apporte "
                "pas de méthode nouvelle à comparer entre les banques."
            ),
        },
    }
    what = what_changed_for_display(change)
    assert "calendrier" in what.lower()
    assert "fragments" not in what.lower()
    assert "passages" not in what.lower()
    assert what.startswith("Le rapport courant")


def test_excel_uses_bank_subject_instead_of_period_aliases() -> None:
    payload = {
        "bank_code": "td",
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "source_text_t1": "Les facteurs existants sont décrits.",
                        "source_text_t2": (
                            "L’incapacité à atteindre les cibles financières est ajoutée."
                        ),
                        "change_summary": (
                            "Le T2 ajoute l’incapacité à atteindre les cibles financières "
                            "parmi les facteurs pouvant créer un écart par rapport aux attentes "
                            "des investisseurs et des analystes."
                        ),
                        "genai_triage": {
                            "is_relevant": True,
                            "nouvelle_idee": True,
                            "impact_level": "MAJEUR",
                            "relevance_reason": (
                                "Le T2 ajoute l’incapacité à atteindre les cibles financières "
                                "parmi les facteurs pouvant créer un écart par rapport aux "
                                "attentes des investisseurs et des analystes. "
                                "Cette précision permet de comparer les facteurs susceptibles "
                                "d’accentuer les écarts aux attentes entre les banques."
                            ),
                        },
                    }
                ],
            }
        ],
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]
    summaries = {
        str(ws.cell(row, _column(ws, "Ce qui change")).value or "")
        for row in range(2, ws.max_row + 1)
    }
    justifications = {
        str(ws.cell(row, _column(ws, "Justification de pertinence (IA)")).value or "")
        for row in range(2, ws.max_row + 1)
    }

    expected = (
        "TD ajoute l’incapacité à atteindre les cibles financières parmi les "
        "facteurs pouvant créer un écart par rapport aux attentes des investisseurs "
        "et des analystes."
    )
    assert summaries == {expected}
    assert justifications == {
        (
            "Cette précision permet de comparer les facteurs susceptibles "
            "d’accentuer les écarts aux attentes entre les banques."
        )
    }
    assert all("T1" not in value and "T2" not in value for value in summaries | justifications)


def test_excel_prefers_structured_units_and_does_not_duplicate_bmo_fact() -> None:
    payload = {
        "bank_code": "bmo",
        "section_comparisons": [
            {
                "section_key": "gestion_risques",
                "section_title": "Gestion des risques",
                "all_block_comparisons": [
                    {
                        "diff_type": "modified",
                        "source_text_t1": "BMO Harris Bank N.A. est mentionnée.",
                        "source_text_t2": "BMO Bank N.A. est mentionnée.",
                        "genai_triage": {
                            "is_relevant": True,
                            "themes_amf": ["GOUVERNANCE_RISQUES"],
                            "impact_level": "MINEUR",
                            "nouvelle_idee": False,
                            "changement_constate": (
                                "Le rapport courant remplace BMO Harris Bank N.A. "
                                "par BMO Bank N.A."
                            ),
                            "signification_metier": (
                                "Cette mise à jour clarifie la dénomination juridique "
                                "utilisée."
                            ),
                            "comparaison_interbanques": (
                                "Elle permet de comparer les entités juridiques visées "
                                "par les banques."
                            ),
                            "limite_interpretation": (
                                "La divulgation ne démontre aucun changement de pratique."
                            ),
                            "motif_non_pertinence": "",
                            "relevance_reason": (
                                "RAISON LEGACY qui ne doit jamais être exportée."
                            ),
                        },
                    }
                ],
            }
        ],
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]
    what_values = {
        str(ws.cell(row, _column(ws, "Ce qui change")).value or "")
        for row in range(2, ws.max_row + 1)
    }
    relevance_values = {
        str(
            ws.cell(
                row,
                _column(ws, "Justification de pertinence (IA)"),
            ).value
            or ""
        )
        for row in range(2, ws.max_row + 1)
    }

    fact = "BMO remplace BMO Harris Bank N.A. par BMO Bank N.A."
    expected_relevance = (
        "Cette mise à jour clarifie la dénomination juridique utilisée. "
        "Elle permet de comparer les entités juridiques visées par les banques. "
        "La divulgation ne démontre aucun changement de pratique."
    )
    assert what_values == {fact}
    assert relevance_values == {expected_relevance}
    assert all(fact not in value for value in relevance_values)
    assert all("LEGACY" not in value for value in what_values | relevance_values)


def test_excel_exports_structured_non_relevance_reason_without_factual_copy() -> None:
    payload = {
        "bank_code": "bmo",
        "section_comparisons": [
            {
                "section_key": "gestion_capital",
                "section_title": "Gestion du capital",
                "all_block_comparisons": [
                    {
                        "diff_type": "added",
                        "source_text_t2": (
                            "La dénomination BMO Bank N.A. est désormais utilisée."
                        ),
                        "genai_triage": {
                            "is_relevant": False,
                            "themes_amf": [],
                            "impact_level": "MINEUR",
                            "nouvelle_idee": False,
                            "exclusion_reason": "reformulation_mineure",
                            "changement_constate": (
                                "BMO actualise la dénomination BMO Bank N.A. "
                                "dans sa divulgation."
                            ),
                            "signification_metier": "",
                            "comparaison_interbanques": "",
                            "limite_interpretation": "",
                            "motif_non_pertinence": (
                                "Cette actualisation rédactionnelle ne révèle aucune "
                                "nouvelle pratique de gestion des fonds propres."
                            ),
                            "relevance_reason": (
                                "RAISON LEGACY qui ne doit pas remplacer le motif."
                            ),
                        },
                    }
                ],
            }
        ],
    }

    raw = generate_text_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Analyse complète"]
    what = str(ws.cell(2, _column(ws, "Ce qui change")).value or "")
    justification = str(
        ws.cell(
            2,
            _column(ws, "Justification de pertinence (IA)"),
        ).value
        or ""
    )

    assert what == (
        "BMO actualise la dénomination BMO Bank N.A. dans sa divulgation."
    )
    assert justification == (
        "Cette actualisation rédactionnelle ne révèle aucune nouvelle pratique "
        "de gestion des fonds propres."
    )
    assert what not in justification
    assert "LEGACY" not in justification
