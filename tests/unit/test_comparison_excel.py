from __future__ import annotations

import io

from openpyxl import load_workbook

from vigilance.comparison_excel import generate_comparison_excel


def test_comparison_excel_prefers_amf_justification() -> None:
    payload = {
        "pair_comparisons": [
            {
                "previous_table": {
                    "section": "gestion_risques",
                    "title": "Risque technologique",
                    "page": 10,
                },
                "current_table": {
                    "section": "gestion_risques",
                    "title": "Risque technologique",
                    "page": 12,
                },
                "genai_triage": {
                    "nouvelle_idee": True,
                    "nouvelle_idee_justification": (
                        "OUI - Nouvel élément à surveiller : Oui.\n\n"
                        "Sujet détecté : Cybersécurité, risque émergent.\n\n"
                        "Ce qui change : Le T2 ajoute une mention cyber absente du T1.\n\n"
                        "Pertinence métier : La justification AMF explique le lien "
                        "avec la vigie bancaire et la comparabilité entre pairs.\n\n"
                        "Point de surveillance : Le point à retenir est le nouveau signal cyber."
                    ),
                },
                "technical_diff": {
                    "indicators_added": [
                        {
                            "value": "Cyberrisque",
                            "analyst_assessment": {
                                "relevance_level": 1,
                                "justification": "Ancienne justification technique.",
                            },
                        }
                    ]
                },
            }
        ]
    }

    raw = generate_comparison_excel(payload, output_path=None)
    workbook = load_workbook(io.BytesIO(raw))
    ws = workbook["Changements détectés"]

    assert "Sujet détecté : Cybersécurité" in ws["I2"].value
    assert "Ancienne justification technique" not in ws["I2"].value
