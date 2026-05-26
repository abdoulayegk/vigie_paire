from __future__ import annotations

import io

from openpyxl import load_workbook

from vigilance.review_export import (
    EXPERT_EXCEL_COLUMNS,
    EXPERT_EXCEL_SHEET_REVIEW,
    EXPERT_EXCEL_SHEET_SUMMARY,
    generate_validation_csv,
    generate_validation_excel,
    generate_validation_txt,
)
from vigilance.review_models import ReviewItem


def test_generate_validation_csv_includes_comment_in_summary_without_schema_change() -> None:
    item = ReviewItem(
        change_id="chg-1",
        change_type="added",
        indicator="Nouvel indicateur",
        section="capital_management",
        page_t2=12,
        review_status="approved",
        comment="Verifier l'impact superviseur",
        edited_value="Nouvel indicateur ajuste",
        confidence=0.91,
    )

    csv_text = generate_validation_csv(
        [item],
        {
            "bank_code": "bnc",
            "quarter_from": "Q1-2025",
            "quarter_to": "Q2-2025",
            "year": 2025,
        },
    )

    assert "Commentaire analyste: Verifier l'impact superviseur." in csv_text
    assert "Valeur editee: Nouvel indicateur ajuste." in csv_text
    assert "validation_finale" in csv_text


def test_generate_validation_excel_generates_expert_workbook() -> None:
    indicator_item = ReviewItem(
        change_id="ind-1",
        change_type="modified",
        indicator="Ratio CET1",
        section="capital_management",
        table_name="Structure des fonds propres",
        page_t1=33,
        page_t2=38,
        review_status="approved",
        comment="Variation cohérente",
        confidence=0.91,
        indicators=[
            {
                "name": "Ratio CET1",
                "type": "modified",
                "from": "12,4 %",
                "to": "12,8 %",
                "review_status": "approved",
            }
        ],
        genai_analysis={"relevance": "NOUVELLE_DIVULGATION"},
    )
    footnote_item = ReviewItem(
        change_id="fn-1",
        change_type="footnote",
        indicator="+1 note(s)",
        section="risk_management",
        table_name="Table risque",
        page_t1=3,
        page_t2=4,
        review_status="pending",
        comment="A revoir",
        confidence=0.75,
        item_type="footnote",
        indicators=[
            {
                "name": "[1] Ancien texte -> Nouveau texte",
                "type": "modified",
                "review_status": "pending",
            }
        ],
    )
    removed_item = ReviewItem(
        change_id="tbl-rem-1",
        change_type="table_removed",
        indicator="Tableau entier retiré",
        section="risk_management",
        table_name="Table supprimée",
        page_t1=41,
        review_status="rejected",
        comment="A reclassifier",
    )

    excel_bytes = generate_validation_excel(
        [indicator_item, footnote_item, removed_item],
        {
            "bank_code": "bnc",
            "quarter_from": "Q1-2025",
            "quarter_to": "Q2-2025",
            "year": 2025,
            "tables_matched": 24,
            "tables_added": [{"table_id": "tbl_cur_1"}],
            "tables_removed": [{"table_id": "tbl_prev_1"}],
        },
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    assert wb.sheetnames == [EXPERT_EXCEL_SHEET_SUMMARY, EXPERT_EXCEL_SHEET_REVIEW]

    ws_review = wb[EXPERT_EXCEL_SHEET_REVIEW]
    headers = [
        ws_review.cell(row=1, column=i).value for i in range(1, ws_review.max_column + 1)
    ]
    assert headers == EXPERT_EXCEL_COLUMNS
    assert "pertinence_genai" not in headers
    assert "niveau_risque_genai" not in headers
    assert "change_id" not in headers
    assert "confidence" not in headers

    assert ws_review.freeze_panes == "A2"
    assert ws_review.auto_filter.ref == ws_review.dimensions
    assert ws_review.max_row == 4

    row_2 = [ws_review.cell(row=2, column=i).value for i in range(1, ws_review.max_column + 1)]
    assert row_2[0] == "Gestion du capital"
    assert row_2[1] == "Structure des fonds propres"
    assert row_2[4] == "33"
    assert row_2[5] == "38"
    assert row_2[6] == "12,4 % → 12,8 %"
    assert row_2[8] == "Non"
    assert row_2[9] == "Validé"
    assert row_2[10] == "Variation cohérente"

    row_3 = [ws_review.cell(row=3, column=i).value for i in range(1, ws_review.max_column + 1)]
    assert row_3[2] == "Note"
    assert row_3[8] == "Non"
    assert row_3[9] == "En attente"

    row_4 = [ws_review.cell(row=4, column=i).value for i in range(1, ws_review.max_column + 1)]
    assert row_4[3] == "retiré"
    assert row_4[4] == "41"
    assert row_4[5] is None
    assert row_4[6] == "Tableau entier retiré"
    assert row_4[8] == "Non"
    assert row_4[9] == "Rejeté"

    ws_summary = wb[EXPERT_EXCEL_SHEET_SUMMARY]
    summary_pairs = {
        str(ws_summary.cell(row=i, column=1).value): ws_summary.cell(row=i, column=2).value
        for i in range(2, 14)
        if ws_summary.cell(row=i, column=1).value not in ("", None)
    }
    assert summary_pairs["Banque"] == "BNC"
    assert summary_pairs["Trimestre comparé"] == "T2 2025 vs T1 2025"
    assert summary_pairs["Nombre total de changements"] == 3
    assert summary_pairs["Validés"] == 1
    assert summary_pairs["Rejetés"] == 1
    assert summary_pairs["En attente"] == 1
    assert summary_pairs["Tableaux appariés"] == 24
    assert summary_pairs["Vrais ajouts"] == 1
    assert summary_pairs["Vrais retraits"] == 1
    assert summary_pairs["Notes modifiées"] == 1

    section_header_row = 15
    section_headers = [
        ws_summary.cell(row=section_header_row, column=i).value for i in range(1, 6)
    ]
    assert section_headers == [
        "Section",
        "Nombre de changements",
        "Validés",
        "Rejetés",
        "En attente",
    ]


def test_validation_excel_prefers_amf_v2_analyst_justification() -> None:
    analyst_note = (
        "OUI - le T2 ajoute un indicateur TLAC absent du T1 dans le tableau "
        "des fonds propres. Cette information change la divulgation prudentielle "
        "et introduit un nouvel axe de lecture.\n\n"
        "Cette nouveaute est pertinente pour la vigie parce qu'elle touche les "
        "ratios reglementaires et les exigences de capital applicables aux "
        "banques canadiennes. Elle ne doit pas etre remplacee par un simple "
        "resume technique.\n\n"
        "L'analyste devrait verifier la methode de calcul, comparer la "
        "presentation avec les pairs et suivre si d'autres tableaux reprennent "
        "la meme exigence."
    )
    item = ReviewItem(
        change_id="ind-amf-v2",
        change_type="modified",
        indicator="Ratio TLAC",
        section="capital_management",
        table_name="Structure des fonds propres",
        page_t1=33,
        page_t2=38,
        indicators=[
            {
                "name": "Ratio TLAC",
                "type": "modified",
                "from": "ancien libelle",
                "to": "nouveau libelle",
                "analyst_assessment": {
                    "nouvelle_idee_justification": analyst_note,
                    "justification": "Ancien resume court a ne pas utiliser.",
                },
            }
        ],
        genai_analysis={
            "nouvelle_idee": True,
            "nouvelle_idee_justification": "Justification table-level moins prioritaire.",
            "justification": "Ancien resume table-level.",
        },
    )

    excel_bytes = generate_validation_excel([item], {"bank_code": "bnc"})
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws_review = wb[EXPERT_EXCEL_SHEET_REVIEW]

    assert ws_review.cell(row=2, column=8).value == analyst_note
    assert ws_review.cell(row=2, column=9).value == "Oui"


def test_generate_validation_txt_generates_readable_report() -> None:
    item = ReviewItem(
        change_id="ind-1",
        change_type="modified",
        indicator="Ratio CET1",
        section="capital_management",
        table_name="Structure des fonds propres",
        page_t1=33,
        page_t2=38,
        review_status="approved",
        comment="Variation cohérente",
        indicators=[
            {
                "name": "Ratio CET1",
                "type": "modified",
                "from": "12,4 %",
                "to": "12,8 %",
                "review_status": "approved",
            }
        ],
        genai_analysis={"relevance": "NOUVELLE_DIVULGATION"},
    )

    txt = generate_validation_txt(
        [item],
        {
            "bank_code": "bnc",
            "quarter_from": "Q1-2025",
            "quarter_to": "Q2-2025",
        },
    )

    assert "REVUE EXPERT" in txt
    assert "Banque : BNC" in txt
    assert "Trimestre comparé : T2 2025 vs T1 2025" in txt
    assert "SECTION : Gestion du capital" in txt
    assert "Tableau : Structure des fonds propres" in txt
    assert "Nouvelle idée : Non" in txt
    assert "Validation expert : Validé" in txt
    assert "Commentaire expert : Variation cohérente" in txt
    assert "change_id" not in txt
    assert "pertinence_genai" not in txt
    assert "niveau_risque_genai" not in txt
