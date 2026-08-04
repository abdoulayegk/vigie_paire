"""Export des résultats de comparaison texte vers un classeur Excel analyste."""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any

from vigie.support.i18n.fr import impact_label_fr, sanitize_analyst_french
from vigie.support.vigie_columns import build_text_vigie_display_row

logger = logging.getLogger(__name__)

_SECTION_DISPLAY: dict[str, str] = {
    "gestion_capital": "Gestion du capital",
    "gestion_risques": "Gestion des risques",
    "gestion_reglementation": "Faits nouveaux en matière de réglementation",
}

_IMPACT_SORT_ORDER: dict[str, int] = {"MAJEUR": 0, "MODERE": 1, "MINEUR": 2}

_CATEGORY_SORT_ORDER: dict[str, int] = {
    "REGLEMENTAIRE": 0,
    "RISQUE": 1,
    "CAPITAL": 2,
    "STRUCTURE": 3,
    "NON_PERTINENT": 4,
    "INCONNU": 5,
}

EXCEL_COLUMNS = [
    "Texte exact du trimestre courant",
    "Texte exact du trimestre précédent",
    "Catégorie principale",
    "Étiquettes secondaires",
    "Section du rapport",
    "Sous-section",
    "Type d'élément",
    "Type de changement",
    "Ce qui change",
    "Nouvelle idée à surveiller ?",
    "Justification de pertinence (IA)",
    "Priorité / impact",
    "Page du texte courant",
    "Page du texte précédent",
    "Statut analyste",
    "Note analyste",
    "Validé le",
]

_VALIDATION_STATUS_FR: dict[str, str] = {
    "approved": "Validé",
    "rejected": "Rejeté",
    "skipped": "Ignoré",
    "pending": "En attente",
    "": "En attente",
}

_CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(value: Any) -> Any:
    """Nettoie les chaînes avant écriture dans openpyxl."""
    if isinstance(value, str):
        return _CONTROL_CHAR_RE.sub("", value)
    return value


# ---------------------------------------------------------------------------
# Tri
# ---------------------------------------------------------------------------

def _row_sort_key(row: dict[str, Any]) -> tuple:
    """Cle de tri analyste : pertinence -> impact -> nouvelle idee."""
    return (
        0 if row.get("is_relevant") else 1,
        _IMPACT_SORT_ORDER.get(str(row.get("impact_level") or "").upper(), 99),
        0 if row.get("nouvelle_idee_bool") else 1,
        _CATEGORY_SORT_ORDER.get(str(row.get("category") or "").upper(), 5),
        str(row.get("section_title") or ""),
        str(row.get("diff_type") or ""),
    )


# ---------------------------------------------------------------------------
# Couleurs par niveau d'impact
# ---------------------------------------------------------------------------

_FILL_NOUVELLE_IDEE = "D6E4F0"   # bleu clair — nouvelle idée
_FILL_MAJEUR        = "FADADD"   # rouge clair
_FILL_MODERE        = "FDEBD0"   # orange clair
_FILL_MINEUR_REL    = "FEF9E7"   # jaune très pâle — MINEUR mais pertinent
_FILL_MINEUR        = "FFFFFF"   # blanc — non pertinent ou MINEUR standard


def _row_fill_color(row: dict[str, Any]) -> str | None:
    """Retourne la couleur de remplissage Excel selon nouvelle_idee et impact."""
    if row.get("nouvelle_idee_bool"):
        return _FILL_NOUVELLE_IDEE
    level = str(row.get("impact_level") or "").upper()
    if level == "MAJEUR":
        return _FILL_MAJEUR
    if level == "MODERE":
        return _FILL_MODERE
    if level == "MINEUR" and row.get("is_relevant"):
        return _FILL_MINEUR_REL
    return None


def _published_change_types(change: dict[str, Any]) -> tuple[str, ...]:
    """Mapper un diff technique vers les seuls types visibles à l'analyste.

    Une modification textuelle contient simultanément un retrait et un ajout.
    Les deux lignes gardent les mêmes preuves T1/T2 afin de rendre le
    remplacement auditable, sans introduire ``Modification`` comme quatrième
    type métier. Un déplacement confirmé n'est pas une vigie publiable.
    """
    if str(change.get("alignment_decision") or "").lower() == "moved_text":
        return ()

    diff_type = str(change.get("diff_type") or "").lower()
    if diff_type == "added":
        return ("Ajout",)
    if diff_type == "removed":
        return ("Suppression",)
    if diff_type == "renamed":
        return ("Renommage",)
    if diff_type == "modified":
        return ("Suppression", "Ajout")
    return ()


def _published_change_summary(change_type: str, display_summary: str) -> str:
    """Conserver la phrase métier; le type possède déjà sa propre colonne."""
    _ = change_type
    return display_summary


# ---------------------------------------------------------------------------
# Collecte des lignes
# ---------------------------------------------------------------------------

def _collect_rows(text_comparison: dict[str, Any]) -> list[dict[str, Any]]:
    """Extrait et aplatit toutes les lignes de changement depuis ``text_comparison.json``."""
    rows: list[dict[str, Any]] = []
    bank_code = str(text_comparison.get("bank_code") or "").strip()
    for section_comp in text_comparison.get("section_comparisons", []):
        section_key = section_comp.get("section_key", "")
        section_title = (
            section_comp.get("section_title")
            or _SECTION_DISPLAY.get(section_key, section_key)
        )
        for block_comp in section_comp.get("all_block_comparisons", []):
            if block_comp.get("diff_type") == "unchanged":
                continue
            published_types = _published_change_types(block_comp)
            if not published_types:
                continue

            triage = block_comp.get("genai_triage") or {}
            evidence_t1 = block_comp.get("evidence_t1") or {}
            evidence_t2 = block_comp.get("evidence_t2") or {}
            page_t1 = ", ".join(str(p) for p in (evidence_t1.get("pages") or []) if p)
            page_t2 = ", ".join(str(p) for p in (evidence_t2.get("pages") or []) if p)

            display = build_text_vigie_display_row(
                block_comp,
                section_title=section_title,
                bank_code=bank_code,
            )
            # La justification exportée contient seulement l'analyse métier
            # (ou le motif de non-pertinence), jamais le constat déjà publié
            # dans la colonne « Ce qui change ».
            justification = display["business_relevance"]
            analyst_review = block_comp.get("_analyst_review") or {}
            analyst_status = str(analyst_review.get("status") or "").strip().lower()
            nouvelle_idee = display["nouvelle_idee_label"]
            analyst_comment = ""
            if analyst_status == "rejected":
                nouvelle_idee = "Non"
                analyst_comment = str(analyst_review.get("comment") or "").strip()
            elif analyst_status == "approved":
                analyst_comment = str(analyst_review.get("comment") or "").strip()

            base_row = {
                "change_id": block_comp.get("change_id", ""),
                "section_key": section_key,
                "section_title": section_title,
                "subsection": display["subsection"],
                "page_t1": page_t1,
                "page_t2": page_t2,
                # Les deux colonnes de texte gardent la source du pipeline,
                # sans résumé ni reformulation générée.
                "source_text_t1": str(block_comp.get("source_text_t1") or ""),
                "source_text_t2": str(block_comp.get("source_text_t2") or ""),
                "impact_level": str(triage.get("impact_level") or "MINEUR").upper(),
                "category": sanitize_analyst_french(str(display["category"] or "")),
                "secondary_labels": sanitize_analyst_french(
                    str(display["secondary_labels"] or "")
                ),
                "is_relevant": bool(triage.get("is_relevant", False)),
                "nouvelle_idee_bool": nouvelle_idee == "Oui",
                "nouvelle_idee": nouvelle_idee,
                "justification": sanitize_analyst_french(justification),
                "analyst_status": _VALIDATION_STATUS_FR.get(analyst_status, analyst_status),
                "commentaire_analyste": analyst_comment,
                "validated_at": str(analyst_review.get("at") or ""),
            }
            for published_type in published_types:
                rows.append(
                    {
                        **base_row,
                        "diff_type": published_type,
                        "what_changed": sanitize_analyst_french(
                            _published_change_summary(
                                published_type,
                                display["what_changed"],
                            )
                        ),
                    }
                )

    rows.sort(key=_row_sort_key)
    return rows


# ---------------------------------------------------------------------------
# Export principal
# ---------------------------------------------------------------------------

def generate_text_comparison_excel(
    text_comparison: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path | bytes:
    """Génère le classeur Excel analyste à partir de text_comparison.json.

    Affiche tous les changements détectés hors ``unchanged``, triés par
    nouvelle idée puis par priorité. Les dates pures, variations chiffrées et
    reformulations restent visibles pour décision analyste.
    L'analyste valide chaque ligne via la colonne Statut.

    Args:
        text_comparison: Dictionnaire text_comparison.json complet.
        output_path: Chemin de sortie .xlsx. Si None, retourne les bytes.

    Returns:
        Path du fichier créé, ou bytes si output_path est None.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Analyse complète"

    # ---------- styles ----------
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="top", wrap_text=True)

    # ---------- en-têtes ----------
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---------- données ----------
    rows = _collect_rows(text_comparison)

    for row_num, row in enumerate(rows, start=2):
        fill_hex = _row_fill_color(row)
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid") if fill_hex else None

        row_data = [
            row["source_text_t2"],
            row["source_text_t1"],
            row["category"],
            row["secondary_labels"],
            row["section_title"],
            row["subsection"],
            "Texte",
            row["diff_type"],
            row["what_changed"],
            row["nouvelle_idee"],
            row["justification"],
            impact_label_fr(str(row["impact_level"] or "MINEUR")),
            row["page_t2"],
            row["page_t1"],
            row["analyst_status"],
            row["commentaire_analyste"],
            row["validated_at"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_excel_safe(value))
            cell.alignment = cell_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

    last_row = len(rows) + 1

    # ---------- largeurs de colonnes ----------
    col_widths = {
        1: 70,   # Texte exact courant
        2: 70,   # Texte exact précédent
        3: 38,   # Catégorie principale
        4: 42,   # Étiquettes secondaires
        5: 30,   # Section
        6: 35,   # Sous-section
        7: 18,   # Type d'élément
        8: 18,   # Type de changement
        9: 70,   # Ce qui change
        10: 20,  # Nouvelle idée
        11: 70,  # Justification
        12: 16,  # Impact
        13: 16,  # Page courante
        14: 16,  # Page précédente
        15: 18,  # Statut analyste
        16: 30,  # Note analyste
        17: 22,  # Validé le
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{max(last_row, 1)}"

    # ---------- sauvegarde ----------
    if output_path is None:
        buf = io.BytesIO()
        wb.save(buf)
        logger.info("text_comparison_excel: %d changements → BytesIO", len(rows))
        return buf.getvalue()

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("text_comparison_excel: %d changements → %s", len(rows), out)
    return out
