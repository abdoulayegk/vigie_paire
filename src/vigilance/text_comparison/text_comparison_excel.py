"""Export des résultats de comparaison texte vers un classeur Excel analyste."""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any

from vigilance.text_comparison.justification import build_text_triage_justification

logger = logging.getLogger(__name__)

_SECTION_DISPLAY: dict[str, str] = {
    "gestion_capital": "Gestion du capital",
    "gestion_risques": "Gestion des risques",
    "gestion_reglementation": "Faits nouveaux en matière de réglementation",
}

_IMPACT_SORT_ORDER: dict[str, int] = {"MAJEUR": 0, "MODERE": 1, "MINEUR": 2}

_CATEGORY_SORT_ORDER: dict[str, int] = {
    "REGLEMENTAIRE": 0,
    "RISQUE": 1,
    "CAPITAL": 2,
    "STRUCTURE": 3,
    "NON_PERTINENT": 4,
    "INCONNU": 5,
}

EXCEL_COLUMNS = [
    "Titre",
    "Sous-section",
    "Page T1",
    "Page T2",
    "Type de changement",
    "Texte exact T1",
    "Texte exact T2",
    "Nouvelle idée ?",
    "Justification IA",
    "Commentaire analyste",
]

_VIGIE_TOPICS: tuple[dict[str, Any], ...] = (
    {
        "tag": "appetit_risque",
        "label": "Appétit pour le risque",
        "objective": "Vérifier les éléments divulgués dans la section appétit pour le risque.",
        "keywords": (
            "appétit pour le risque",
            "appetit pour le risque",
            "limite de risque",
            "limites de risque",
            "risk appetite",
        ),
    },
    {
        "tag": "risques_esg",
        "label": "Risque ESG",
        "objective": "Vérifier ce qui est inclus dans la section Risques ESG.",
        "keywords": ("esg", "environnemental", "social", "durabilité", "durabilite", "droits de la personne"),
    },
    {
        "tag": "edtf_ifi",
        "label": "EDTF et importance systémique",
        "objective": "Vérifier les textes des pairs par rapport à l'EDTF et aux banques d'importance systémique (IFIS).",
        "keywords": (
            "edtf",
            "importance systémique",
            "importance systemique",
            "ifis",
            "ifi",
            "bism",
            "bisn",
            "banque d'importance systémique",
            "banque d importance systemique",
        ),
    },
    {
        "tag": "ro_calcul_fonds_propres",
        "label": "RO (calcul des fonds propres)",
        "objective": "Vérifier le niveau de détails que donnent les pairs quant au calcul.",
        "keywords": (
            "fonds propres",
            "capital réglementaire",
            "capital reglementaire",
            "actifs pondérés",
            "actifs ponderes",
            "ratio cet1",
            "tlac",
            "approche ni",
        ),
    },
    {
        "tag": "climat_credit",
        "label": "Impact du risque climatique sur le risque de crédit",
        "objective": "Vérifier si les pairs abordent le sujet des impacts climatiques sur le risque de crédit.",
        "keywords": (
            "risque climatique",
            "risques climatiques",
            "changements climatiques",
            "climat",
            "risque de crédit",
            "risque de credit",
            "b-15",
        ),
    },
    {
        "tag": "ia",
        "label": "Intelligence artificielle (IA)",
        "objective": "Vérifier le contenu du texte sur l'IA.",
        "keywords": ("intelligence artificielle", " ia ", "ia générative", "ia generative", "générative", "generative"),
    },
    {
        "tag": "politique_monetaire",
        "label": "Politique monétaire",
        "objective": "Vérifier si les politiques monétaires sont traitées dans une section dédiée ou intégrées dans une autre section.",
        "keywords": ("politique monétaire", "politique monetaire", "politiques monétaires", "politiques monetaires", "taux d'intérêt"),
    },
    {
        "tag": "endettement_menages",
        "label": "Endettement des ménages",
        "objective": "Vérifier que les pairs abordent le sujet et le contenu du texte.",
        "keywords": ("endettement des ménages", "endettement des menages", "ménages", "menages", "habitation", "hypothécaire"),
    },
    {
        "tag": "capacite_recruter",
        "label": "Capacité à recruter",
        "objective": "Vérifier que les pairs abordent le sujet et le contenu du texte.",
        "keywords": ("recruter", "recrutement", "rétention", "retention", "talents", "main-d'œuvre", "main d'oeuvre"),
    },
)

_TOPIC_BY_TAG = {str(topic["tag"]): topic for topic in _VIGIE_TOPICS}
_TOPIC_BY_LABEL = {str(topic["label"]): topic for topic in _VIGIE_TOPICS}

_DIFF_TYPE_FR: dict[str, str] = {
    "added": "Ajout",
    "removed": "Suppression",
    "modified": "Modification",
    "renamed": "Renommage",
    "unchanged": "Inchangé",
}

# ---------------------------------------------------------------------------
# Heuristiques d'exclusion (dates pures et reformulations strictes)
# ---------------------------------------------------------------------------

_DATE_UPDATE_RE = re.compile(
    r"(mise à jour|changement|modification|passé).{0,40}(date|période|trimestre|semestre|référence de clôture)|"
    r"(date|période de référence).{0,20}(mise à jour|changé|modifié|passé)|"
    r"\b(janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\b"
    r".{0,25}→.{0,25}"
    r"\b(janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\b|"
    r"^(la )?date (de référence|de clôture|des échéances) (a été|est) (mis|modif|chang)",
    flags=re.IGNORECASE,
)

_REFORMULATION_RE = re.compile(
    r"\b(légère|simple|même|pure).{0,20}(reformulation|reformulé|reformulée)\b|"
    r"\breformulation.{0,20}(légère|sans changement|sans nouveau fond)\b|"
    r"\bmême (idée|information|sens|contenu).{0,30}(reformul|formulat différente)\b",
    flags=re.IGNORECASE,
)

_CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(value: Any) -> Any:
    """Nettoie les chaînes avant écriture dans openpyxl."""
    if isinstance(value, str):
        return _CONTROL_CHAR_RE.sub("", value)
    return value


def _normalize_topic_text(value: str) -> str:
    """Normalise un texte pour les recherches de topics de vigie."""
    normalized = value.lower()
    normalized = (
        normalized.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("ë", "e")
        .replace("à", "a")
        .replace("â", "a")
        .replace("ù", "u")
        .replace("û", "u")
        .replace("î", "i")
        .replace("ï", "i")
        .replace("ô", "o")
        .replace("ç", "c")
        .replace("œ", "oe")
    )
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _shorten(value: str, limit: int = 260) -> str:
    """Retourne une chaîne courte adaptée à une cellule de synthèse."""
    clean = re.sub(r"\s+", " ", value or "").strip()
    if len(clean) <= limit:
        return clean
    return clean[: max(0, limit - 1)].rstrip() + "…"


def _page_values(value: str) -> list[int]:
    """Parse les pages stockées sous forme de texte '1, 2'."""
    pages: list[int] = []
    for item in re.findall(r"\d+", value or ""):
        try:
            pages.append(int(item))
        except ValueError:
            continue
    return pages


def _page_range(values: list[int]) -> str:
    """Formate une liste de pages en plage lisible."""
    unique = sorted(set(values))
    if not unique:
        return ""
    if len(unique) == 1:
        return f"p. {unique[0]}"
    if unique[-1] - unique[0] + 1 == len(unique):
        return f"p. {unique[0]}-{unique[-1]}"
    return "p. " + ", ".join(str(page) for page in unique)


def _join_unique(values: list[str], *, limit: int | None = None) -> str:
    """Joint les valeurs uniques non vides dans l'ordre d'apparition."""
    seen: list[str] = []
    for value in values:
        clean = str(value or "").strip()
        if clean and clean not in seen:
            seen.append(clean)
        if limit is not None and len(seen) >= limit:
            break
    return "; ".join(seen)


def _topic_labels_for_change(change: dict[str, Any]) -> list[str]:
    """Retourne les topics de vigie associés à un changement."""
    labels: list[str] = []
    triage = change.get("genai_triage") or {}

    for objective in list(change.get("objective_matches") or []) + list(
        triage.get("objective_matches") or []
    ):
        label = str(objective.get("label") or "").strip()
        tag = str(objective.get("tag") or "").strip()
        topic = _TOPIC_BY_TAG.get(tag) or _TOPIC_BY_LABEL.get(label)
        resolved = str((topic or {}).get("label") or label).strip()
        if resolved and resolved not in labels:
            labels.append(resolved)

    for tag in list(change.get("objective_tags") or []) + list(triage.get("objective_tags") or []):
        topic = _TOPIC_BY_TAG.get(str(tag))
        label = str((topic or {}).get("label") or "").strip()
        if label and label not in labels:
            labels.append(label)

    search_text = " ".join(
        str(change.get(key) or "")
        for key in (
            "current_subsection_heading",
            "previous_subsection_heading",
            "subsection_heading",
            "canonical_topic",
            "chunk_topic",
            "change_summary",
            "source_text_t1",
            "source_text_t2",
        )
    )
    normalized = f" {_normalize_topic_text(search_text)} "
    for topic in _VIGIE_TOPICS:
        label = str(topic["label"])
        if label in labels:
            continue
        for keyword in topic["keywords"]:
            keyword_norm = f" {_normalize_topic_text(str(keyword))} "
            if keyword_norm.strip() and keyword_norm in normalized:
                labels.append(label)
                break

    return labels


def _is_pure_date_update(change: dict[str, Any]) -> bool:
    """Vrai uniquement pour les mises à jour de dates sans autre contenu."""
    triage = change.get("genai_triage") or {}
    if str(triage.get("impact_level") or "").upper() != "MINEUR":
        return False
    if str(triage.get("category") or "").upper() != "NON_PERTINENT":
        return False
    summary = change.get("change_summary") or ""
    return bool(_DATE_UPDATE_RE.search(summary))


def _is_pure_reformulation(change: dict[str, Any]) -> bool:
    """Vrai uniquement pour les reformulations strictement identiques."""
    triage = change.get("genai_triage") or {}
    if str(triage.get("impact_level") or "").upper() != "MINEUR":
        return False
    if str(triage.get("category") or "").upper() != "NON_PERTINENT":
        return False
    summary = change.get("change_summary") or ""
    return bool(_REFORMULATION_RE.search(summary))


def _should_exclude(change: dict[str, Any]) -> bool:
    """Indique si le changement doit être exclu de l'export (date ou reformulation)."""
    return _is_pure_date_update(change) or _is_pure_reformulation(change)


# ---------------------------------------------------------------------------
# Extraction de la sous-section
# ---------------------------------------------------------------------------

def _subsection_label(change: dict[str, Any]) -> str:
    """Retourne le nom lisible de la sous-section."""
    heading = (
        change.get("current_subsection_heading")
        or change.get("subsection_heading")
        or change.get("previous_subsection_heading")
        or ""
    )
    if heading and heading not in ("__intro__", "full", ""):
        return heading
    # Fallback: extraire du change_id
    change_id = change.get("change_id") or ""
    section_key = change.get("section_key") or ""
    prefix = section_key + "_"
    if change_id.startswith(prefix):
        slug = re.sub(r"_change_\d+$", "", change_id[len(prefix):])
        if slug and slug != "full":
            return slug.replace("_", " ").strip()
    return ""


# ---------------------------------------------------------------------------
# Tri
# ---------------------------------------------------------------------------

def _row_sort_key(row: dict[str, Any]) -> tuple:
    """Cle de tri analyste : pertinence -> impact -> nouvelle idee."""
    return (
        0 if row.get("is_relevant") else 1,
        _IMPACT_SORT_ORDER.get(str(row.get("impact_level") or "").upper(), 99),
        0 if row.get("nouvelle_idee_bool") else 1,
        _CATEGORY_SORT_ORDER.get(str(row.get("category") or "").upper(), 5),
        str(row.get("section_title") or ""),
        str(row.get("diff_type") or ""),
    )


# ---------------------------------------------------------------------------
# Couleurs par niveau d'impact
# ---------------------------------------------------------------------------

_FILL_NOUVELLE_IDEE = "D6E4F0"   # bleu clair — nouvelle idée
_FILL_MAJEUR        = "FADADD"   # rouge clair
_FILL_MODERE        = "FDEBD0"   # orange clair
_FILL_MINEUR_REL    = "FEF9E7"   # jaune très pâle — MINEUR mais pertinent
_FILL_MINEUR        = "FFFFFF"   # blanc — non pertinent ou MINEUR standard


def _row_fill_color(row: dict[str, Any]) -> str | None:
    """Retourne la couleur de remplissage Excel selon nouvelle_idee et impact."""
    if row.get("nouvelle_idee_bool"):
        return _FILL_NOUVELLE_IDEE
    level = str(row.get("impact_level") or "").upper()
    if level == "MAJEUR":
        return _FILL_MAJEUR
    if level == "MODERE":
        return _FILL_MODERE
    if level == "MINEUR" and row.get("is_relevant"):
        return _FILL_MINEUR_REL
    return None


# ---------------------------------------------------------------------------
# Collecte des lignes
# ---------------------------------------------------------------------------

def _collect_rows(text_comparison: dict[str, Any]) -> list[dict[str, Any]]:
    """Extrait et aplatit toutes les lignes de changement depuis ``text_comparison.json``."""
    rows: list[dict[str, Any]] = []
    for section_comp in text_comparison.get("section_comparisons", []):
        section_key = section_comp.get("section_key", "")
        section_title = (
            section_comp.get("section_title")
            or _SECTION_DISPLAY.get(section_key, section_key)
        )
        source_changes = section_comp.get("all_observation_comparisons")
        if not isinstance(source_changes, list):
            source_changes = section_comp.get("all_block_comparisons", [])
        for block_comp in source_changes:
            if not isinstance(block_comp, dict):
                continue
            if block_comp.get("diff_type") == "unchanged":
                continue

            triage = block_comp.get("genai_triage") or {}
            evidence_t1 = block_comp.get("evidence_t1") or {}
            evidence_t2 = block_comp.get("evidence_t2") or {}
            page_t1 = ", ".join(str(p) for p in (evidence_t1.get("pages") or []) if p)
            page_t2 = ", ".join(str(p) for p in (evidence_t2.get("pages") or []) if p)

            justification = build_text_triage_justification(block_comp)
            analyst_review = block_comp.get("_analyst_review") or {}
            analyst_status = str(analyst_review.get("status") or "").strip().lower()
            ai_nouvelle_idee = bool(triage.get("nouvelle_idee", False))
            nouvelle_idee = "Oui" if ai_nouvelle_idee else "Non"
            analyst_comment = ""
            if analyst_status == "rejected":
                nouvelle_idee = "Non"
                analyst_comment = str(analyst_review.get("comment") or "").strip()
            elif analyst_status == "approved":
                analyst_comment = str(analyst_review.get("comment") or "").strip()

            rows.append(
                {
                    "change_id": block_comp.get("change_id", ""),
                    "section_key": section_key,
                    "section_title": section_title,
                    "subsection": _subsection_label(block_comp),
                    "chunk_topic": str(block_comp.get("chunk_topic") or "").strip(),
                    "canonical_topic": str(block_comp.get("canonical_topic") or "").strip(),
                    "change_summary": str(block_comp.get("change_summary") or "").strip(),
                    "page_t1": page_t1,
                    "page_t2": page_t2,
                    "evidence_t1": str(evidence_t1.get("snippet") or "").strip(),
                    "evidence_t2": str(evidence_t2.get("snippet") or "").strip(),
                    "source_text_t1": (block_comp.get("source_text_t1") or "").strip(),
                    "source_text_t2": (block_comp.get("source_text_t2") or "").strip(),
                    "diff_type": block_comp.get("diff_type", ""),
                    "impact_level": str(triage.get("impact_level") or "MINEUR").upper(),
                    "category": str(triage.get("category") or "INCONNU").upper(),
                    "is_relevant": bool(triage.get("is_relevant", False)),
                    "action_requise": str(triage.get("action_requise") or "aucune").strip(),
                    "exclusion_reason": str(triage.get("exclusion_reason") or "").strip(),
                    "nouvelle_idee_bool": nouvelle_idee == "Oui",
                    "nouvelle_idee": nouvelle_idee,
                    "justification": justification,
                    "commentaire_analyste": analyst_comment,
                    "topic_labels": _topic_labels_for_change(block_comp),
                }
            )

    rows.sort(key=_row_sort_key)
    return rows


def _rows_for_topic(rows: list[dict[str, Any]], topic_label: str) -> list[dict[str, Any]]:
    """Filtre les lignes associées à un topic de vigie."""
    return [row for row in rows if topic_label in (row.get("topic_labels") or [])]


def _topic_level(rows: list[dict[str, Any]]) -> str:
    """Calcule un niveau métier synthétique pour un topic."""
    if not rows:
        return ""
    impacts = {str(row.get("impact_level") or "").upper() for row in rows}
    if any(row.get("nouvelle_idee_bool") for row in rows) or "MAJEUR" in impacts:
        return "Important"
    if any(row.get("is_relevant") for row in rows) or "MODERE" in impacts:
        return "Moyen"
    return "Faible"


def _topic_verdict(rows: list[dict[str, Any]]) -> str:
    """Retourne le verdict de couverture pour un topic."""
    if not rows:
        return "Non détecté"
    if any(row.get("is_relevant") or row.get("nouvelle_idee_bool") for row in rows):
        return "Traité"
    return "Traité - mineur"


def _topic_pages(rows: list[dict[str, Any]], key: str) -> str:
    """Retourne la plage de pages pour un topic."""
    pages: list[int] = []
    for row in rows:
        pages.extend(_page_values(str(row.get(key) or "")))
    return _page_range(pages)


def _topic_change_ids(rows: list[dict[str, Any]], *, limit: int | None = None) -> str:
    """Retourne les change_id uniques d'un groupe."""
    return _join_unique([str(row.get("change_id") or "") for row in rows], limit=limit)


def _topic_summary_text(rows: list[dict[str, Any]]) -> str:
    """Construit une synthèse déterministe courte pour un topic."""
    if not rows:
        return "Aucun changement textuel associé à ce topic dans les résultats comparés."
    sections = _join_unique([str(row.get("section_title") or "") for row in rows], limit=3)
    subsections = _join_unique([str(row.get("subsection") or "") for row in rows], limit=5)
    intro = f"Traité dans {sections}." if sections else "Traité dans les résultats comparés."
    if subsections:
        intro += f" Sous-sections principales: {subsections}."
    summaries = _join_unique(
        [
            _shorten(str(row.get("change_summary") or row.get("chunk_topic") or ""), 180)
            for row in rows
            if row.get("change_summary") or row.get("chunk_topic")
        ],
        limit=3,
    )
    if summaries:
        intro += f" Points observés: {summaries}."
    return intro


def _build_vigie_summary_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    """Construit les lignes de l'onglet Synthèse Vigie."""
    output: list[list[Any]] = []
    for topic in _VIGIE_TOPICS:
        label = str(topic["label"])
        topic_rows = _rows_for_topic(rows, label)
        output.append(
            [
                label,
                _topic_verdict(topic_rows),
                _topic_pages(topic_rows, "page_t2"),
                _topic_pages(topic_rows, "page_t1"),
                _topic_summary_text(topic_rows),
                _join_unique(
                    [
                        _shorten(str(row.get("change_summary") or row.get("chunk_topic") or ""), 160)
                        for row in topic_rows
                        if row.get("change_summary") or row.get("chunk_topic")
                    ],
                    limit=4,
                ),
                _topic_level(topic_rows),
                _topic_change_ids(topic_rows, limit=10),
                str(topic["objective"]),
            ]
        )
    return output


def _build_vigie_detail_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    """Construit les lignes consolidées par topic/sous-thème."""
    grouped: dict[tuple[str, str, str, str], list[dict[str, Any]]] = {}
    for row in rows:
        for topic_label in row.get("topic_labels") or []:
            key = (
                str(topic_label),
                str(row.get("section_title") or ""),
                str(row.get("subsection") or ""),
                str(row.get("chunk_topic") or row.get("canonical_topic") or ""),
            )
            grouped.setdefault(key, []).append(row)

    output: list[list[Any]] = []
    for (topic_label, section, subsection, chunk_topic), topic_rows in sorted(grouped.items()):
        output.append(
            [
                topic_label,
                chunk_topic or subsection,
                section,
                subsection,
                _topic_pages(topic_rows, "page_t2"),
                _topic_pages(topic_rows, "page_t1"),
                _join_unique(
                    [
                        _shorten(str(row.get("change_summary") or ""), 180)
                        for row in topic_rows
                        if row.get("change_summary")
                    ],
                    limit=3,
                ),
                _topic_level(topic_rows),
                _topic_verdict(topic_rows),
                _topic_change_ids(topic_rows),
            ]
        )
    return output


def _build_vigie_evidence_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    """Construit les lignes de preuves, une ligne par topic/change_id."""
    output: list[list[Any]] = []
    for row in rows:
        topic_labels = row.get("topic_labels") or ["Hors topics vigie"]
        for topic_label in topic_labels:
            output.append(
                [
                    row.get("change_id"),
                    topic_label,
                    row.get("section_title"),
                    row.get("subsection"),
                    row.get("page_t2"),
                    row.get("page_t1"),
                    row.get("chunk_topic"),
                    row.get("change_summary"),
                    row.get("impact_level"),
                    "Oui" if row.get("is_relevant") else "Non",
                    row.get("exclusion_reason"),
                    _shorten(str(row.get("evidence_t2") or row.get("source_text_t2") or ""), 500),
                    _shorten(str(row.get("evidence_t1") or row.get("source_text_t1") or ""), 500),
                ]
            )
    return output


def _build_vigie_review_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    """Construit les lignes de l'onglet À revoir."""
    output: list[list[Any]] = []
    for topic in _VIGIE_TOPICS:
        label = str(topic["label"])
        topic_rows = _rows_for_topic(rows, label)
        if not topic_rows:
            output.append(
                [
                    label,
                    "Non détecté",
                    "Aucun changement textuel associé à ce topic dans les résultats comparés.",
                    "Vérifier la couverture dans l'extraction complète si le sujet est attendu.",
                    "",
                ]
            )
            continue
        if not any(row.get("is_relevant") or row.get("nouvelle_idee_bool") for row in topic_rows):
            output.append(
                [
                    label,
                    "À confirmer",
                    "Sujet repéré seulement dans des changements mineurs ou non pertinents selon l'IA.",
                    "Confirmer si le sujet doit rester dans la vigie métier.",
                    _topic_change_ids(topic_rows, limit=10),
                ]
            )
    return output


# ---------------------------------------------------------------------------
# Export principal
# ---------------------------------------------------------------------------

def _generate_text_comparison_workbook(
    text_comparison: dict[str, Any],
    *,
    include_vigie_sheets: bool,
) -> Any:
    """Construit le classeur Excel, avec ou sans les onglets vigie."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    rows = _collect_rows(text_comparison)

    # ---------- styles ----------
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="top", wrap_text=True)

    def write_table(
        sheet: Any,
        headers: list[str],
        data_rows: list[list[Any]],
        widths: dict[int, int],
    ) -> None:
        for col_idx, col_name in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_num, data_row in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row_num, column=col_idx, value=_excel_safe(value))
                cell.alignment = cell_align
                cell.border = thin_border
        for col_idx, width in widths.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A2"
        last_row = max(len(data_rows) + 1, 1)
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    if include_vigie_sheets:
        # ---------- onglets vigie ----------
        ws_summary = wb.active
        ws_summary.title = "Synthèse Vigie"
        write_table(
            ws_summary,
            [
                "Topic vigie",
                "Verdict",
                "Pages T2",
                "Pages T1",
                "Synthèse analyste",
                "Changements clés",
                "Niveau",
                "Preuves",
                "Objectif de vigie",
            ],
            _build_vigie_summary_rows(rows),
            {1: 34, 2: 18, 3: 12, 4: 12, 5: 70, 6: 60, 7: 14, 8: 45, 9: 70},
        )

        ws_detail = wb.create_sheet("Détail par topic")
        write_table(
            ws_detail,
            [
                "Topic vigie",
                "Sous-thème",
                "Section",
                "Sous-section",
                "Pages T2",
                "Pages T1",
                "Résumé consolidé",
                "Niveau",
                "Verdict",
                "change_ids",
            ],
            _build_vigie_detail_rows(rows),
            {1: 34, 2: 35, 3: 28, 4: 38, 5: 12, 6: 12, 7: 70, 8: 14, 9: 18, 10: 55},
        )

        ws_evidence = wb.create_sheet("Preuves")
        write_table(
            ws_evidence,
            [
                "change_id",
                "Topic vigie",
                "Section",
                "Sous-section",
                "Page T2",
                "Page T1",
                "chunk_topic",
                "change_summary",
                "Impact",
                "Pertinent IA ?",
                "Raison exclusion",
                "Extrait T2",
                "Extrait T1",
            ],
            _build_vigie_evidence_rows(rows),
            {
                1: 42,
                2: 34,
                3: 28,
                4: 38,
                5: 10,
                6: 10,
                7: 35,
                8: 70,
                9: 12,
                10: 14,
                11: 24,
                12: 70,
                13: 70,
            },
        )

        ws_review = wb.create_sheet("À revoir")
        write_table(
            ws_review,
            ["Topic vigie", "Statut", "Raison", "Action analyste", "Preuves"],
            _build_vigie_review_rows(rows),
            {1: 34, 2: 18, 3: 70, 4: 55, 5: 55},
        )

        ws = wb.create_sheet("Analyse complète")
    else:
        ws = wb.active
        ws.title = "Analyse complète"

    # ---------- en-têtes ----------
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---------- données ----------
    for row_num, row in enumerate(rows, start=2):
        fill_hex = _row_fill_color(row)
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid") if fill_hex else None

        row_data = [
            row["section_title"],
            row["subsection"],
            row["page_t1"],
            row["page_t2"],
            _DIFF_TYPE_FR.get(str(row["diff_type"]).lower(), row["diff_type"]),
            row["source_text_t1"],
            row["source_text_t2"],
            row["nouvelle_idee"],
            row["justification"],
            row["commentaire_analyste"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_excel_safe(value))
            cell.alignment = cell_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

    last_row = len(rows) + 1

    # ---------- largeurs de colonnes ----------
    col_widths = {
        1: 30,   # Titre
        2: 35,   # Sous-section
        3: 10,   # Page T1
        4: 10,   # Page T2
        5: 18,   # Type de changement
        6: 70,   # Texte exact T1
        7: 70,   # Texte exact T2
        8: 14,   # Nouvelle idée ?
        9: 70,   # Justification IA
        10: 25,  # Commentaire analyste
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{max(last_row, 1)}"

    return wb, len(rows)


def _save_workbook(
    workbook: Any,
    output_path: str | Path | None,
    *,
    row_count: int,
    export_name: str,
) -> Path | bytes:
    """Sauvegarde un classeur ou retourne ses bytes."""
    if output_path is None:
        buf = io.BytesIO()
        workbook.save(buf)
        logger.info("%s: %d changements → BytesIO", export_name, row_count)
        return buf.getvalue()

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(out))
    logger.info("%s: %d changements → %s", export_name, row_count, out)
    return out


def generate_text_comparison_excel(
    text_comparison: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path | bytes:
    """Génère le classeur Excel analyste standard à partir de text_comparison.json.

    Ce fichier conserve le format historique avec l'onglet ``Analyse complète``.
    """
    workbook, row_count = _generate_text_comparison_workbook(
        text_comparison,
        include_vigie_sheets=False,
    )
    return _save_workbook(
        workbook,
        output_path,
        row_count=row_count,
        export_name="text_comparison_excel",
    )


def generate_text_vigie_excel(
    text_comparison: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path | bytes:
    """Génère le classeur Excel de vigie métier séparé."""
    workbook, row_count = _generate_text_comparison_workbook(
        text_comparison,
        include_vigie_sheets=True,
    )
    return _save_workbook(
        workbook,
        output_path,
        row_count=row_count,
        export_name="text_vigie_excel",
    )
