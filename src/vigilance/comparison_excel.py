"""Export des résultats de comparaison de tableaux vers un classeur Excel analyste.

Ce module génère un fichier Excel à partir de comparison.json, avec une ligne
par changement d'indicateur ou de note de bas de page, incluant la
justification GPT (analyst_assessment) pour chaque élément.

Le fichier est sauvegardé à côté de comparison.json dans le répertoire de sortie.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from vigilance.i18n.fr import sanitize_analyst_french

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colonnes du classeur
# ---------------------------------------------------------------------------

EXCEL_COLUMNS = [
    "Section",
    "Tableau",
    "Page (trimestre précédent)",
    "Page (trimestre courant)",
    "Type d'élément",
    "Type de changement",
    "Libellé (trimestre précédent)",
    "Libellé (trimestre courant)",
    "Justification IA",
    "Nouvelle idée ?",
    "Commentaire analyste",
    "Statut validation",
    "Validé le",
]

_VALIDATION_STATUS_FR: dict[str, str] = {
    "approved": "Validé",
    "rejected": "Rejeté",
    "skipped": "Ignoré",
    "pending": "En attente",
    "": "En attente",
}


def _review_fields(item: dict[str, Any] | None) -> dict[str, str]:
    """Extrait les champs de revue analyste a partir d'un item technical_diff."""
    review = (item or {}).get("_analyst_review") if isinstance(item, dict) else None
    if not isinstance(review, dict):
        return {"notes": "", "status": "", "by": "", "at": ""}
    status_raw = str(review.get("status") or "").strip().lower()
    return {
        "notes": str(review.get("notes") or "").strip(),
        "status": _VALIDATION_STATUS_FR.get(status_raw, status_raw),
        "by": str(review.get("by") or "").strip(),
        "at": str(review.get("at") or "").strip(),
    }

_CHANGE_TYPE_FR: dict[str, str] = {
    "added": "Ajout",
    "removed": "Suppression",
    "renamed": "Renommage",
    "modified": "Modification",
}

_ELEMENT_TYPE_FR: dict[str, str] = {
    "indicator": "Indicateur",
    "footnote": "Note de bas de page",
    "table": "Tableau",
}

_SECTION_FR: dict[str, str] = {
    "gestion_capital": "Gestion du capital",
    "capital_management": "Gestion du capital",
    "gestion_risques": "Gestion des risques",
    "risk_management": "Gestion des risques",
    "gestion_reglementation": "Faits nouveaux en matière de réglementation",
    "regulatory_updates": "Réglementation",
    "reglementation": "Réglementation",
}

_RELEVANCE_SORT: dict[int, int] = {1: 0, 2: 1, 3: 2}

# ---------------------------------------------------------------------------
# Couleurs par niveau
# ---------------------------------------------------------------------------

_FILL_CRITIQUE = "FADADD"   # rouge clair
_FILL_ELEVE = "FDEBD0"     # orange clair
_FILL_FAIBLE = "FFFFFF"     # blanc


def _row_fill_color(row: dict[str, Any]) -> str | None:
    """Retourne la couleur de remplissage Excel selon le niveau de pertinence."""
    level = row.get("relevance_level_raw")
    if level == 1:
        return _FILL_CRITIQUE
    if level == 2:
        return _FILL_ELEVE
    return None


def _section_label(section: str) -> str:
    """Retourne le libellé français d'une section, ou la clé brute si inconnue."""
    return _SECTION_FR.get(str(section).strip().lower(), section)


# ---------------------------------------------------------------------------
# Collecte des lignes depuis comparison.json
# ---------------------------------------------------------------------------

def _collect_rows(comparison: dict[str, Any]) -> list[dict[str, Any]]:
    """Extrait toutes les lignes de changement depuis comparison.json."""
    rows: list[dict[str, Any]] = []

    # --- Paires appariées ---
    for pair in comparison.get("pair_comparisons", []):
        prev_table = pair.get("previous_table") or {}
        cur_table = pair.get("current_table") or {}
        section = _section_label(
            cur_table.get("section") or prev_table.get("section") or ""
        )
        title = cur_table.get("title") or prev_table.get("title") or "(sans titre)"
        page_t1 = prev_table.get("page", "")
        page_t2 = cur_table.get("page", "")
        diff = pair.get("technical_diff") or {}

        # Triage global du tableau
        triage = pair.get("genai_triage") or {}
        table_justification = str(triage.get("nouvelle_idee_justification") or "").strip()
        table_is_relevant = bool(triage.get("is_relevant", False))
        table_category = str(triage.get("category") or "").upper()

        # Indicateurs ajoutés
        for ind in diff.get("indicators_added") or []:
            if not isinstance(ind, dict):
                continue
            assessment = ind.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["indicator"],
                "change_type": _CHANGE_TYPE_FR["added"],
                "label_t1": "",
                "label_t2": str(ind.get("value") or ""),
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(ind),
            })

        # Indicateurs supprimés
        for ind in diff.get("indicators_removed") or []:
            if not isinstance(ind, dict):
                continue
            assessment = ind.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["indicator"],
                "change_type": _CHANGE_TYPE_FR["removed"],
                "label_t1": str(ind.get("value") or ""),
                "label_t2": "",
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(ind),
            })

        # Indicateurs renommés
        for ind in diff.get("indicators_renamed") or []:
            if not isinstance(ind, dict):
                continue
            assessment = ind.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["indicator"],
                "change_type": _CHANGE_TYPE_FR["renamed"],
                "label_t1": str(ind.get("previous") or ""),
                "label_t2": str(ind.get("current") or ""),
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(ind),
            })

        # Notes ajoutées
        for fn in diff.get("footnotes_added") or []:
            if not isinstance(fn, dict):
                continue
            assessment = fn.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["footnote"],
                "change_type": _CHANGE_TYPE_FR["added"],
                "label_t1": "",
                "label_t2": str(fn.get("text") or ""),
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(fn),
            })

        # Notes supprimées
        for fn in diff.get("footnotes_removed") or []:
            if not isinstance(fn, dict):
                continue
            assessment = fn.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["footnote"],
                "change_type": _CHANGE_TYPE_FR["removed"],
                "label_t1": str(fn.get("text") or ""),
                "label_t2": "",
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(fn),
            })

        # Notes modifiées (renommées)
        for fn in diff.get("footnotes_renamed") or []:
            if not isinstance(fn, dict):
                continue
            assessment = fn.get("analyst_assessment") or {}
            justification = _best_analyst_justification(assessment, table_justification)
            relevance_level = assessment.get("relevance_level")
            rows.append({
                "section": section,
                "title": title,
                "page_t1": str(page_t1) if page_t1 else "",
                "page_t2": str(page_t2) if page_t2 else "",
                "element_type": _ELEMENT_TYPE_FR["footnote"],
                "change_type": _CHANGE_TYPE_FR.get("modified", "Modification"),
                "label_t1": str(fn.get("previous_text") or ""),
                "label_t2": str(fn.get("current_text") or ""),
                "justification": justification,
                "relevance_level_raw": relevance_level,
                "nouvelle_idee": _is_nouvelle_idee_label(triage),
                "review": _review_fields(fn),
            })

    # --- Tableaux ajoutés ---
    matching = comparison.get("matching") or {}
    for tbl in matching.get("tables_added") or []:
        if not isinstance(tbl, dict):
            continue
        triage = tbl.get("genai_triage") or {}
        justification = sanitize_analyst_french(
            str(triage.get("nouvelle_idee_justification") or "").strip()
        )
        category = str(triage.get("category") or "").upper()
        is_relevant = bool(triage.get("is_relevant", False))
        rows.append({
            "section": _section_label(tbl.get("section") or ""),
            "title": str(tbl.get("title") or "(sans titre)"),
            "page_t1": "",
            "page_t2": str(tbl.get("page") or ""),
            "element_type": _ELEMENT_TYPE_FR["table"],
            "change_type": _CHANGE_TYPE_FR["added"],
            "label_t1": "",
            "label_t2": str(tbl.get("title") or ""),
            "justification": justification or "Nouveau tableau détecté dans le trimestre courant.",
            "relevance_level_raw": None,
            "nouvelle_idee": _is_nouvelle_idee_label(triage),
            "review": _review_fields(tbl),
        })

    # --- Tableaux supprimés ---
    for tbl in matching.get("tables_removed") or []:
        if not isinstance(tbl, dict):
            continue
        triage = tbl.get("genai_triage") or {}
        justification = sanitize_analyst_french(
            str(triage.get("nouvelle_idee_justification") or "").strip()
        )
        category = str(triage.get("category") or "").upper()
        is_relevant = bool(triage.get("is_relevant", False))
        rows.append({
            "section": _section_label(tbl.get("section") or ""),
            "title": str(tbl.get("title") or "(sans titre)"),
            "page_t1": str(tbl.get("page") or ""),
            "page_t2": "",
            "element_type": _ELEMENT_TYPE_FR["table"],
            "change_type": _CHANGE_TYPE_FR["removed"],
            "label_t1": str(tbl.get("title") or ""),
            "label_t2": "",
            "justification": justification or "Tableau supprimé du trimestre courant.",
            "relevance_level_raw": None,
            "nouvelle_idee": _is_nouvelle_idee_label(triage),
            "review": _review_fields(tbl),
        })

    # Tri : pertinence critique en premier, puis section, puis tableau
    rows.sort(key=lambda r: (
        _RELEVANCE_SORT.get(r.get("relevance_level_raw") or 99, 99),
        r.get("section", ""),
        r.get("title", ""),
        r.get("change_type", ""),
    ))

    return rows


def _is_nouvelle_idee_label(triage: dict[str, Any]) -> str:
    """Retourne 'Oui' / 'Non' selon la décision GPT (champ ``nouvelle_idee``)."""
    return "Oui" if bool(triage.get("nouvelle_idee", False)) else "Non"


def _best_analyst_justification(
    assessment: dict[str, Any],
    table_justification: str,
) -> str:
    """Privilégie la justification AMF enrichie, puis le fallback technique."""
    amf_justification = table_justification.strip()
    if amf_justification:
        return sanitize_analyst_french(amf_justification)
    return sanitize_analyst_french(str(assessment.get("justification") or "").strip())


# ---------------------------------------------------------------------------
# Export principal
# ---------------------------------------------------------------------------

def generate_comparison_excel(
    comparison: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path | bytes:
    """Génère le classeur Excel analyste à partir de comparison.json.

    Chaque ligne correspond à un changement d'indicateur ou de note de bas
    de page, avec la justification IA (GPT) associée.

    Args:
        comparison: Dictionnaire comparison.json complet.
        output_path: Chemin de sortie .xlsx. Si None, retourne les bytes.

    Returns:
        Path du fichier créé, ou bytes si output_path est None.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Changements détectés"

    # ---------- styles ----------
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="top", wrap_text=True)

    # ---------- en-têtes ----------
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---------- données ----------
    rows = _collect_rows(comparison)

    for row_num, row in enumerate(rows, start=2):
        fill_hex = _row_fill_color(row)
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid") if fill_hex else None

        review = row.get("review") or {}
        row_data = [
            row["section"],
            row["title"],
            row["page_t1"],
            row["page_t2"],
            row["element_type"],
            row["change_type"],
            row["label_t1"],
            row["label_t2"],
            row["justification"],
            row["nouvelle_idee"],
            review.get("notes", ""),
            review.get("status", ""),
            review.get("at", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

    last_row = len(rows) + 1

    # ---------- largeurs de colonnes ----------
    col_widths = {
        1: 30,   # Section
        2: 45,   # Tableau
        3: 10,   # Page (trimestre précédent)
        4: 10,   # Page (trimestre courant)
        5: 22,   # Type d'élément
        6: 18,   # Type de changement
        7: 50,   # Libellé (trimestre précédent)
        8: 50,   # Libellé (trimestre courant)
        9: 70,   # Justification IA
        10: 16,  # Nouvelle idée ?
        11: 30,  # Commentaire analyste
        12: 16,  # Statut validation
        13: 22,  # Validé le
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{max(last_row, 1)}"

    # ---------- sauvegarde ----------
    if output_path is None:
        buf = io.BytesIO()
        wb.save(buf)
        logger.info("comparison_excel: %d changements → BytesIO", len(rows))
        return buf.getvalue()

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("comparison_excel: %d changements → %s", len(rows), out)
    return out
