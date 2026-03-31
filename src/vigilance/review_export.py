"""Export helpers for review items and analyst-facing expert Excel exports."""

from __future__ import annotations

import csv
import io
import json
from collections import Counter, defaultdict
from datetime import datetime
from typing import Any, Iterator

from vigilance.review_models import ReviewItem

CSV_SCHEMA_VERSION = "csv_review_v1"
CSV_SEPARATOR = ";"
CSV_BOM = "\ufeff"

# Legacy CSV validation schema kept for flat exports.
VALIDATION_CSV_COLUMNS = [
    "banque",
    "trimestre",
    "tableau",
    "section",
    "type_élément",
    "type_changement",
    "page_precedente",
    "page_courante",
    "indicateur_precedent",
    "indicateur_courant",
    "résumé_automatique",
    "validation_finale",
    "notes_analyste",
    "date_validation",
    "nouvelle_divulgation",
    "pertinence_genai",
    "niveau_risque_genai",
]

EXPERT_EXCEL_SHEET_SUMMARY = "Synthèse"
EXPERT_EXCEL_SHEET_REVIEW = "Revue expert"
EXPERT_EXCEL_COLUMNS = [
    "Banque",
    "Trimestre comparé",
    "Section",
    "Tableau",
    "Type d’élément",
    "Type de changement",
    "Page précédente",
    "Page courante",
    "Valeur / libellé précédent",
    "Valeur / libellé courant",
    "Résumé métier",
    "Nouvelle idée ?",
    "Validation expert",
    "Commentaire expert",
    "Date de validation",
]

# Mapping type_changement anglais -> francais
_TYPE_CHANGEMENT_MAP = {
    "added": "ajouté",
    "removed": "retiré",
    "renamed": "renommé",
    "table_added": "ajouté",
    "table_removed": "retiré",
    "modified": "modifié",
    "modifie": "modifié",
    "displaced": "déplacé",
    "uncertain": "incertain",
    "needs_review": "incertain",
    "structure_change": "fusion",
    "footnote": "note modifiee",
}

# Match methods consideres suspects
_SUSPECT_MATCH_METHODS = frozenset(
    {
        "split_merge_rescue",
        "single_rescue",
        "uncertain_competition",
        "unknown_section_penalized",
        "low_label_overlap_reject",
    }
)

# Mapping section (code interne) -> libellé français
_SECTION_FR = {
    "gestion_capital": "Gestion du capital",
    "capital_management": "Gestion du capital",
    "gestion_risques": "Gestion des risques",
    "risk_management": "Gestion des risques",
    "gestion_reglementation": "Réglementation",
    "regulatory_updates": "Réglementation",
    "reglementation": "Réglementation",
}

# Mapping review_status -> validation_finale
_VALIDATION_FINALE_MAP = {
    "pending": "En attente",
    "approved": "Validé",
    "rejected": "Rejeté",
}

_CSV_COLUMNS = [
    "run_id",
    "date_generation",
    "version_schema",
    "bank_code",
    "quarter_from",
    "quarter_to",
    "year",
    "change_id",
    "section",
    "table_name",
    "table_number",
    "table_id_t1",
    "table_id_t2",
    "page_t1",
    "page_t2",
    "table_status",
    "indicator_name",
    "indicator_type",
    "indicator_from",
    "indicator_to",
    "resume_court",
    "confidence",
    "match_method",
    "review_status",
    "idee",
    "sujet",
    "validation_finale",
    "comment",
]


def _sanitize_cell(val: Any) -> str:
    """Nettoie retours ligne pour CSV Excel. Les guillemets/delimiteurs sont geres par csv.DictWriter."""
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return s


def _format_cell(val: Any) -> str:
    """Formate une valeur pour CSV (None, int, float). Pages en entiers si possible."""
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else str(val)
    return _sanitize_cell(val)


def _build_resume_court(
    indicator_type: str,
    indicator_name: str,
    from_val: str,
    to_val: str,
    table_status: str = "",
    suspect: bool = False,
) -> str:
    """Genere une phrase interpretative courte pour le superviseur (FR, 1 phrase max)."""
    if indicator_type == "added":
        base = "Nouvel indicateur dans le trimestre courant."
    elif indicator_type == "removed":
        base = "Indicateur présent au trimestre précédent mais absent au trimestre courant."
    elif indicator_type == "renamed" and (from_val or to_val):
        base = "Renommage probable d'indicateur."
    elif indicator_type == "renamed":
        base = "Renommage probable d'indicateur."
    elif indicator_type == "table_added":
        base = "Nouveau tableau dans le trimestre courant."
    elif indicator_type == "table_removed":
        base = (
            "Tableau présent au trimestre précédent mais absent au trimestre courant."
        )
    elif indicator_type == "modified":
        base = "Modification de tableau détectée."
    elif indicator_type == "uncertain":
        base = "Aucun match clair trouvé — revue manuelle requise."
    elif indicator_type == "structure_change":
        base = "Fusion ou split probable — revue manuelle requise."
    elif indicator_type == "footnote":
        base = "Note de bas de tableau modifiee -- verifier impact reglementaire."
    elif "note" in indicator_type.lower() or table_status == "note":
        base = "Texte de note modifié — vérifier impact réglementaire."
    else:
        base = _sanitize_cell(indicator_name) or "Changement détecté."
    if suspect:
        base = f"{base} Revue manuelle recommandée."
    return base


def _augment_resume_with_review_context(
    resume: str,
    *,
    comment: str = "",
    edited_value: str = "",
) -> str:
    """Append analyst review context without changing the supervisor CSV schema."""
    parts = [str(resume or "").strip()]
    comment_clean = _sanitize_cell(comment)
    edited_clean = _sanitize_cell(edited_value)
    if edited_clean:
        parts.append(f"Valeur editee: {edited_clean}.")
    if comment_clean:
        parts.append(f"Commentaire analyste: {comment_clean}.")
    return " ".join(part for part in parts if part).strip()


def _build_export_context(indicator_result: dict[str, Any] | None) -> dict[str, str]:
    """Extract stable export context from the canonical comparison payload."""
    ir = indicator_result or {}
    meta = ir.get("meta") or {}
    previous = (meta.get("extraction_sources") or {}).get("previous") or {}
    current = (meta.get("extraction_sources") or {}).get("current") or {}
    return {
        "bank_code": _sanitize_cell(ir.get("bank_code", "")),
        "quarter_from": _sanitize_cell(
            ir.get("quarter_from") or ir.get("previous_quarter") or ""
        ),
        "quarter_to": _sanitize_cell(
            ir.get("quarter_to") or ir.get("current_quarter") or ""
        ),
        "year": _format_cell(ir.get("year", "")),
        "compare_path": _sanitize_cell(meta.get("compare_path", "")),
        "generated_at": _sanitize_cell(meta.get("generated_at", "")),
        "previous_source_mode": _sanitize_cell(previous.get("mode", "")),
        "current_source_mode": _sanitize_cell(current.get("mode", "")),
        "previous_tables_path": _sanitize_cell(previous.get("tables_path", "")),
        "current_tables_path": _sanitize_cell(current.get("tables_path", "")),
        "previous_indicators_path": _sanitize_cell(previous.get("indicators_path", "")),
        "current_indicators_path": _sanitize_cell(current.get("indicators_path", "")),
        "previous_footnotes_path": _sanitize_cell(previous.get("footnotes_path", "")),
        "current_footnotes_path": _sanitize_cell(current.get("footnotes_path", "")),
    }


def build_export_metadata(
    indicator_result: dict[str, Any] | None,
    *,
    overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Public helper to build stable export metadata from a comparison payload."""
    payload = dict(_build_export_context(indicator_result))
    if overrides:
        payload.update(dict(overrides))
    return payload


def _to_type_changement(ind_type: str, table_status: str) -> str:
    """Convertit indicator_type + table_status en type_changement francais."""
    if table_status == "structure_change":
        return "fusion"
    return _TYPE_CHANGEMENT_MAP.get(ind_type, "incertain")


def _to_type_element(ind_type: str, item_type: str = "indicator") -> str:
    """Retourne type_élément: tableau, indicateur, note ou texte."""
    if item_type == "footnote" or ind_type == "footnote":
        return "note"
    if ind_type in (
        "table_added",
        "table_removed",
        "modified",
        "uncertain",
        "structure_change",
    ):
        return "tableau"
    if ind_type in ("note", "texte"):
        return ind_type
    return "indicateur"


def _to_suspect(match_method: str) -> str:
    """Convertit match_method en suspect Oui/Non."""
    if not match_method:
        return "Non"
    normalized = str(match_method).strip().lower()
    for suspect in _SUSPECT_MATCH_METHODS:
        if suspect.lower() in normalized:
            return "Oui"
    return "Non"


def _section_to_fr(section: str) -> str:
    """Convertit le code section en libellé français pour le CSV."""
    if not section:
        return ""
    key = str(section).strip().lower()
    return _SECTION_FR.get(key, section)


def _to_validation_finale(review_status: str) -> str:
    """Convertit review_status en validation_finale."""
    return _VALIDATION_FINALE_MAP.get(
        str(review_status or "").strip().lower(),
        "En attente",
    )


def _to_nouvelle_idee(genai_analysis: dict[str, Any] | None) -> str:
    """Return a conservative business-facing novelty flag."""
    relevance = str((genai_analysis or {}).get("relevance", "") or "").strip().upper()
    return "Oui" if relevance == "NOUVELLE_DIVULGATION" else "Non"


def _build_trimestre_label(indicator_result: dict[str, Any] | None) -> str:
    """Build a human-facing quarter comparison label."""
    ir = indicator_result or {}
    ctx = _build_export_context(ir)
    q_from = _sanitize_cell(ctx.get("quarter_from", ""))
    q_to = _sanitize_cell(ctx.get("quarter_to", ""))
    if not q_to and not q_from:
        return ""
    if q_to and q_from:
        return f"{q_to.upper()} vs {q_from.upper()}"
    return q_to.upper() or q_from.upper()


def _iter_expert_excel_rows(
    review_items: list[ReviewItem],
    indicator_result: dict[str, Any] | None,
) -> Iterator[dict[str, str]]:
    """Yield analyst-facing expert Excel rows with one line per business change."""
    ir = indicator_result or {}
    banque = _sanitize_cell(ir.get("bank_code", "")).upper()
    trimestre = _build_trimestre_label(ir)

    for item in review_items:
        base = item.to_dict()
        indicators = list(base.get("indicators", []) or [])
        table_status = str(base.get("table_status", "") or "")
        item_type = str(base.get("item_type", "indicator") or "indicator")
        tableau = _sanitize_cell(base.get("table_name", ""))
        section = _section_to_fr(base.get("section", ""))
        validation = _to_validation_finale(base.get("review_status", ""))
        notes_analyste = _sanitize_cell(base.get("comment", ""))
        date_validation = ""
        ts = str(base.get("review_timestamp", "") or "")
        if len(ts) >= 10 and ts[4] == "-":
            date_validation = f"{ts[8:10]}/{ts[5:7]}/{ts[:4]}"
        nouvelle_idee = _to_nouvelle_idee(base.get("genai_analysis") or {})

        if not indicators:
            ind_type = str(base.get("change_type", "") or "")
            type_elem = _to_type_element(ind_type, item_type)
            type_chg = _to_type_changement(ind_type, table_status)
            precedent = ""
            courant = ""
            display_indicator = _sanitize_cell(base.get("indicator", ""))
            if ind_type in ("table_removed", "removed"):
                precedent = display_indicator
            elif ind_type in ("table_added", "added"):
                courant = display_indicator
            else:
                precedent = display_indicator
                courant = display_indicator

            resume = _build_resume_court(
                ind_type,
                display_indicator,
                precedent,
                courant,
                table_status,
                suspect=False,
            )
            resume = _augment_resume_with_review_context(
                resume,
                comment=notes_analyste,
                edited_value=str(base.get("edited_value", "") or ""),
            )
            yield {
                "Banque": banque,
                "Trimestre comparé": trimestre,
                "Section": section,
                "Tableau": tableau,
                "Type d’élément": type_elem.capitalize(),
                "Type de changement": type_chg,
                "Page précédente": _format_cell(base.get("page_t1")),
                "Page courante": _format_cell(base.get("page_t2")),
                "Valeur / libellé précédent": precedent,
                "Valeur / libellé courant": courant,
                "Résumé métier": resume,
                "Nouvelle idée ?": nouvelle_idee,
                "Validation expert": validation,
                "Commentaire expert": notes_analyste,
                "Date de validation": date_validation,
            }
            continue

        for ind in indicators:
            ind_type = str(ind.get("type", "") or "")
            item_change_type = str(base.get("change_type", "") or "")
            type_elem = _to_type_element(item_change_type or ind_type, item_type)
            type_chg = _to_type_changement(ind_type, table_status)
            ind_name = _sanitize_cell(ind.get("name", ""))
            from_val = _sanitize_cell(ind.get("from", ""))
            to_val = _sanitize_cell(ind.get("to", ""))

            if item_type == "footnote":
                precedent = _sanitize_cell(ind.get("old_text", "")) or ind_name
                courant = _sanitize_cell(ind.get("new_text", "")) or ind_name
            elif ind_type == "added":
                precedent = ""
                courant = ind_name
            elif ind_type == "removed":
                precedent = ind_name
                courant = ""
            elif ind_type == "renamed":
                precedent = from_val or ind_name
                courant = to_val or ind_name
            else:
                precedent = from_val or ind_name
                courant = to_val or ind_name

            resume_type = (
                item_change_type
                if item_change_type in ("table_added", "table_removed")
                else ind_type
            )
            resume = _build_resume_court(
                resume_type,
                ind_name,
                precedent,
                courant,
                table_status,
                suspect=False,
            )
            resume = _augment_resume_with_review_context(
                resume,
                comment=notes_analyste,
                edited_value=str(base.get("edited_value", "") or ""),
            )
            ind_validation = _to_validation_finale(
                ind.get("review_status", base.get("review_status", ""))
            )
            yield {
                "Banque": banque,
                "Trimestre comparé": trimestre,
                "Section": section,
                "Tableau": tableau,
                "Type d’élément": type_elem.capitalize(),
                "Type de changement": type_chg,
                "Page précédente": _format_cell(base.get("page_t1")),
                "Page courante": _format_cell(base.get("page_t2")),
                "Valeur / libellé précédent": precedent,
                "Valeur / libellé courant": courant,
                "Résumé métier": resume,
                "Nouvelle idée ?": nouvelle_idee,
                "Validation expert": ind_validation,
                "Commentaire expert": notes_analyste,
                "Date de validation": date_validation,
            }


def _append_summary_sheet(
    wb: Any,
    rows: list[dict[str, str]],
    indicator_result: dict[str, Any] | None,
) -> None:
    """Create the pilotage summary worksheet for the expert workbook."""
    ir = indicator_result or {}
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl: no active sheet")
    ws.title = EXPERT_EXCEL_SHEET_SUMMARY

    banque = _sanitize_cell(ir.get("bank_code", "")).upper()
    trimestre = _build_trimestre_label(ir)
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    counts = Counter(str(row.get("Validation expert", "") or "") for row in rows)
    sections: dict[str, Counter[str]] = defaultdict(Counter)
    notes_modifiees = 0
    for row in rows:
        section = str(row.get("Section", "") or "")
        validation = str(row.get("Validation expert", "") or "")
        sections[section]["Nombre de changements"] += 1
        sections[section][validation] += 1
        if str(row.get("Type d’élément", "")).strip().lower() == "note":
            notes_modifiees += 1

    tables_added = len(list((ir.get("tables_added") or [])))
    tables_removed = len(list((ir.get("tables_removed") or [])))
    tables_matched = int(ir.get("tables_matched") or ir.get("matched_pairs") or 0)

    ws.append(["Synthèse export expert", ""])
    summary_rows = [
        ("Banque", banque),
        ("Trimestre comparé", trimestre),
        ("Date d’export", exported_at),
        ("", ""),
        ("Nombre total de changements", len(rows)),
        ("Validés", counts.get("Validé", 0)),
        ("Rejetés", counts.get("Rejeté", 0)),
        ("En attente", counts.get("En attente", 0)),
        ("Tableaux appariés", tables_matched),
        ("Vrais ajouts", tables_added),
        ("Vrais retraits", tables_removed),
        ("Notes modifiées", notes_modifiees),
    ]
    for pair in summary_rows:
        ws.append(list(pair))

    ws.append(["", ""])
    ws.append(["Section", "Nombre de changements", "Validés", "Rejetés", "En attente"])
    for section in sorted(sections):
        counter = sections[section]
        ws.append(
            [
                section,
                counter.get("Nombre de changements", 0),
                counter.get("Validé", 0),
                counter.get("Rejeté", 0),
                counter.get("En attente", 0),
            ]
        )


def _style_expert_workbook(wb: Any) -> None:
    """Apply lightweight analyst-friendly formatting."""
    from openpyxl.styles import Font, PatternFill

    review_ws = wb[EXPERT_EXCEL_SHEET_REVIEW]
    summary_ws = wb[EXPERT_EXCEL_SHEET_SUMMARY]

    review_ws.freeze_panes = "A2"
    summary_ws.freeze_panes = "A2"
    review_ws.auto_filter.ref = review_ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in (review_ws, summary_ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    widths = {
        "A": 12,
        "B": 22,
        "C": 22,
        "D": 48,
        "E": 18,
        "F": 18,
        "G": 15,
        "H": 15,
        "I": 28,
        "J": 28,
        "K": 48,
        "L": 16,
        "M": 18,
        "N": 40,
        "O": 18,
    }
    for col, width in widths.items():
        review_ws.column_dimensions[col].width = width
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 22
    summary_ws.column_dimensions["C"].width = 14
    summary_ws.column_dimensions["D"].width = 14
    summary_ws.column_dimensions["E"].width = 14

    status_fill = {
        "Validé": PatternFill("solid", fgColor="C6EFCE"),
        "Rejeté": PatternFill("solid", fgColor="FFC7CE"),
        "En attente": PatternFill("solid", fgColor="FCE4D6"),
    }
    status_col = EXPERT_EXCEL_COLUMNS.index("Validation expert") + 1
    for row_idx in range(2, review_ws.max_row + 1):
        cell = review_ws.cell(row=row_idx, column=status_col)
        fill = status_fill.get(str(cell.value or ""))
        if fill is not None:
            cell.fill = fill


def generate_validation_txt(
    review_items: list[ReviewItem],
    indicator_result: dict[str, Any] | None,
) -> str:
    """Generate a human-readable text report for expert review."""
    rows = list(_iter_expert_excel_rows(review_items, indicator_result))
    ir = indicator_result or {}
    banque = _sanitize_cell(ir.get("bank_code", "")).upper()
    trimestre = _build_trimestre_label(ir)
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    counts = Counter(str(row.get("Validation expert", "") or "") for row in rows)

    lines = [
        "REVUE EXPERT",
        "",
        f"Banque : {banque or '-'}",
        f"Trimestre comparé : {trimestre or '-'}",
        f"Date d’export : {exported_at}",
        "",
        "SYNTHÈSE",
        f"- Nombre total de changements : {len(rows)}",
        f"- Validés : {counts.get('Validé', 0)}",
        f"- Rejetés : {counts.get('Rejeté', 0)}",
        f"- En attente : {counts.get('En attente', 0)}",
        "",
    ]

    grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        grouped[str(row.get("Section", "") or "Sans section")][
            str(row.get("Tableau", "") or "Sans titre")
        ].append(row)

    for section, tables in grouped.items():
        lines.append(f"SECTION : {section}")
        for table_name, table_rows in tables.items():
            lines.append(f"  Tableau : {table_name}")
            for row in table_rows:
                lines.append(
                    f"    - {row.get('Type d’élément', '')} | {row.get('Type de changement', '')}"
                )
                lines.append(
                    f"      Pages : préc. {row.get('Page précédente', '') or '-'} / cour. {row.get('Page courante', '') or '-'}"
                )
                lines.append(
                    f"      Avant : {row.get('Valeur / libellé précédent', '') or '-'}"
                )
                lines.append(
                    f"      Après : {row.get('Valeur / libellé courant', '') or '-'}"
                )
                lines.append(
                    f"      Résumé métier : {row.get('Résumé métier', '') or '-'}"
                )
                lines.append(
                    f"      Nouvelle idée : {row.get('Nouvelle idée ?', '') or 'Non'}"
                )
                lines.append(
                    f"      Validation expert : {row.get('Validation expert', '') or 'En attente'}"
                )
                if row.get("Commentaire expert"):
                    lines.append(
                        f"      Commentaire expert : {row.get('Commentaire expert', '')}"
                    )
                if row.get("Date de validation"):
                    lines.append(
                        f"      Date de validation : {row.get('Date de validation', '')}"
                    )
            lines.append("")
        if lines and lines[-1] != "":
            lines.append("")

    return "\n".join(lines).strip() + "\n"


def _iter_validation_rows(
    review_items: list[ReviewItem],
    indicator_result: dict[str, Any] | None,
) -> Iterator[dict[str, str]]:
    """Itere sur les lignes du schema validation (12 colonnes) pour CSV ou Excel."""
    ir = indicator_result or {}
    ctx = _build_export_context(ir)
    banque = _sanitize_cell(ir.get("bank_code", ""))
    q_from = _sanitize_cell(ctx.get("quarter_from", ""))
    q_to = _sanitize_cell(ctx.get("quarter_to", ""))
    year = _sanitize_cell(ctx.get("year", ""))
    trimestre = f"{q_to.upper()} {year} vs {q_from.upper()}".strip() if q_to else ""

    for item in review_items:
        base = item.to_dict()
        indicators = base.get("indicators", [])
        table_status = str(base.get("table_status", ""))
        confidence_raw = base.get("confidence")
        if confidence_raw is not None:
            try:
                score = round(float(confidence_raw), 2)
            except (TypeError, ValueError):
                score = 0.0
        else:
            score = 0.0
        score_str = f"{score:.2f}"
        match_method = str(base.get("match_method", ""))
        suspect = _to_suspect(match_method)
        validation = _to_validation_finale(base.get("review_status", ""))
        comment = str(base.get("comment", "") or "")
        edited_value = str(base.get("edited_value", "") or "")

        tableau = _sanitize_cell(base.get("table_name", ""))
        notes_analyste = _sanitize_cell(comment)
        ts = str(base.get("review_timestamp", "") or "")
        date_validation = (
            f"{ts[8:10]}/{ts[5:7]}/{ts[:4]}" if len(ts) >= 10 and ts[4] == "-" else ""
        )

        item_type = str(base.get("item_type", "indicator"))
        ga = base.get("genai_analysis") or {}
        pertinence_genai = _sanitize_cell(ga.get("relevance", ""))
        niveau_risque_genai = _sanitize_cell(ga.get("risk_level", ""))
        impact_type_genai = _sanitize_cell(ga.get("impact_type", ""))
        phase_genai = _sanitize_cell(ga.get("project_phase", ""))
        action_genai = _sanitize_cell(ga.get("action_requise", ""))
        ref_regl_genai = _sanitize_cell(ga.get("reference_reglementaire", ""))

        if not pertinence_genai:
            nouvelle_divulgation = "Non analysé"
        elif pertinence_genai.upper() == "NOUVELLE_DIVULGATION":
            nouvelle_divulgation = "Oui"
        else:
            nouvelle_divulgation = "Non"

        if not indicators:
            ind_type = str(base.get("change_type", ""))
            ind_name = _sanitize_cell(base.get("indicator", ""))
            type_elem = _to_type_element(ind_type, item_type)
            type_chg = _to_type_changement(ind_type, table_status)
            resume = _build_resume_court(
                ind_type, ind_name, "", "", table_status, suspect=(suspect == "Oui")
            )
            resume = _augment_resume_with_review_context(
                resume,
                comment=comment,
                edited_value=edited_value,
            )

            if ind_type in ("added", "table_added"):
                ind_t1, ind_t2 = "", ind_name
            elif ind_type in ("removed", "table_removed"):
                ind_t1, ind_t2 = ind_name, ""
            else:
                ind_t1, ind_t2 = ind_name, ind_name

            row = {
                "banque": banque,
                "trimestre": trimestre,
                "tableau": tableau,
                "section": _section_to_fr(base.get("section", "")),
                "type_élément": type_elem,
                "type_changement": type_chg,
                "page_precedente": _format_cell(base.get("page_t1")),
                "page_courante": _format_cell(base.get("page_t2")),
                "indicateur_precedent": ind_t1,
                "indicateur_courant": ind_t2,
                "résumé_automatique": resume,
                "validation_finale": validation,
                "notes_analyste": notes_analyste,
                "date_validation": date_validation,
                "nouvelle_divulgation": nouvelle_divulgation,
                "pertinence_genai": pertinence_genai,
                "niveau_risque_genai": niveau_risque_genai,
            }
            yield row
        else:
            item_change_type = str(base.get("change_type", ""))
            for ind in indicators:
                ind_type = str(ind.get("type", ""))
                ind_name = str(ind.get("name", ""))
                from_val = str(ind.get("from", ""))
                to_val = str(ind.get("to", ""))

                type_elem = _to_type_element(item_change_type or ind_type, item_type)
                type_chg = _to_type_changement(ind_type, table_status)
                resume_type = (
                    item_change_type
                    if item_change_type in ("table_added", "table_removed")
                    else ind_type
                )
                resume = _build_resume_court(
                    resume_type,
                    ind_name,
                    from_val,
                    to_val,
                    table_status,
                    suspect=False,
                )
                resume = _augment_resume_with_review_context(
                    resume,
                    comment=comment,
                    edited_value=edited_value,
                )

                if ind_type == "added":
                    ind_t1, ind_t2 = "", _sanitize_cell(ind_name)
                elif ind_type == "removed":
                    ind_t1, ind_t2 = _sanitize_cell(ind_name), ""
                elif ind_type == "renamed":
                    ind_t1 = (
                        _sanitize_cell(from_val)
                        if from_val
                        else _sanitize_cell(ind_name)
                    )
                    ind_t2 = (
                        _sanitize_cell(to_val) if to_val else _sanitize_cell(ind_name)
                    )
                else:
                    ind_t1, ind_t2 = _sanitize_cell(ind_name), _sanitize_cell(ind_name)

                ind_validation = _to_validation_finale(
                    ind.get("review_status", base.get("review_status", ""))
                )
                row = {
                    "banque": banque,
                    "trimestre": trimestre,
                    "tableau": tableau,
                    "section": _section_to_fr(base.get("section", "")),
                    "type_élément": type_elem,
                    "type_changement": type_chg,
                    "page_precedente": _format_cell(base.get("page_t1")),
                    "page_courante": _format_cell(base.get("page_t2")),
                    "indicateur_precedent": ind_t1,
                    "indicateur_courant": ind_t2,
                    "résumé_automatique": resume,
                    "validation_finale": ind_validation,
                    "notes_analyste": notes_analyste,
                    "date_validation": date_validation,
                    "nouvelle_divulgation": nouvelle_divulgation,
                    "pertinence_genai": pertinence_genai,
                    "niveau_risque_genai": niveau_risque_genai,
                }
                yield row


def generate_validation_csv(
    review_items: list[ReviewItem],
    indicator_result: dict[str, Any] | None,
) -> str:
    """Genere le CSV de validation unique (12 colonnes, Excel FR, conformite bancaire).

    Schema exact:
    banque | section | type_élément | type_changement | page_precedente | page_courante |
    indicateur_precedent | indicateur_courant | résumé_automatique |
    score_confiance | suspect | validation_finale
    """
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=VALIDATION_CSV_COLUMNS,
        delimiter=CSV_SEPARATOR,
        quoting=csv.QUOTE_MINIMAL,
    )
    writer.writeheader()
    for row in _iter_validation_rows(review_items, indicator_result):
        writer.writerow(row)
    return CSV_BOM + buffer.getvalue()


def generate_validation_excel(
    review_items: list[ReviewItem],
    indicator_result: dict[str, Any] | None,
) -> bytes:
    """Generate the expert analyst workbook (.xlsx)."""
    from openpyxl import Workbook

    wb = Workbook()
    expert_rows = list(_iter_expert_excel_rows(review_items, indicator_result))
    _append_summary_sheet(wb, expert_rows, indicator_result)

    ws = wb.create_sheet(EXPERT_EXCEL_SHEET_REVIEW)
    ws.append(list(EXPERT_EXCEL_COLUMNS))
    for row in expert_rows:
        ws.append([row[col] for col in EXPERT_EXCEL_COLUMNS])

    _style_expert_workbook(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_review_items_csv(
    items: list[ReviewItem],
    *,
    metadata: dict[str, Any] | None = None,
    separator: str = CSV_SEPARATOR,
) -> str:
    """Serialize review items to CSV text (schema technique, colonnes en anglais).

    DEPRECATED: Pour l'export superviseur en français, utiliser generate_validation_csv()
    qui produit le schema 12 colonnes (banque, section, type_élément, etc.).
    """
    now = datetime.now()
    meta = metadata or {}
    run_id = str(meta.get("run_id", "") or now.strftime("%Y%m%d_%H%M"))
    date_gen = meta.get("date_generation", "") or now.isoformat(timespec="seconds")
    bank = _sanitize_cell(meta.get("bank_code", ""))
    q_from = _sanitize_cell(meta.get("quarter_from", "t1"))
    q_to = _sanitize_cell(meta.get("quarter_to", "t2"))
    year_val = _format_cell(meta.get("year", ""))

    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=_CSV_COLUMNS,
        delimiter=separator,
        quoting=csv.QUOTE_MINIMAL,
    )
    writer.writeheader()

    for item in items:
        base = item.to_dict()
        indicators = base.get("indicators", [])

        if not indicators:
            row = {
                "run_id": run_id,
                "date_generation": date_gen,
                "version_schema": CSV_SCHEMA_VERSION,
                "bank_code": bank,
                "quarter_from": q_from,
                "quarter_to": q_to,
                "year": year_val,
                "change_id": _sanitize_cell(base.get("change_id", "")),
                "section": _sanitize_cell(base.get("section", "")),
                "table_name": _sanitize_cell(base.get("table_name", "")),
                "table_number": _sanitize_cell(base.get("table_number", "")),
                "table_id_t1": _sanitize_cell(base.get("table_id_t1", "")),
                "table_id_t2": _sanitize_cell(base.get("table_id_t2", "")),
                "page_t1": _format_cell(base.get("page_t1")),
                "page_t2": _format_cell(base.get("page_t2")),
                "table_status": _sanitize_cell(base.get("table_status", "")),
                "indicator_name": _sanitize_cell(base.get("indicator", "")),
                "indicator_type": _sanitize_cell(base.get("change_type", "")),
                "indicator_from": "",
                "indicator_to": "",
                "resume_court": _build_resume_court(
                    base.get("change_type", ""),
                    base.get("indicator", ""),
                    "",
                    "",
                    base.get("table_status", ""),
                ),
                "confidence": _format_cell(base.get("confidence", "")),
                "match_method": _sanitize_cell(base.get("match_method", "")),
                "review_status": _sanitize_cell(base.get("review_status", "")),
                "idee": "",
                "sujet": "",
                "validation_finale": "",
                "comment": _sanitize_cell(base.get("comment", "")),
            }
            writer.writerow(row)
        else:
            for ind in indicators:
                ind_type = str(ind.get("type", ""))
                ind_name = str(ind.get("name", ""))
                from_val = str(ind.get("from", ""))
                to_val = str(ind.get("to", ""))

                row = {
                    "run_id": run_id,
                    "date_generation": date_gen,
                    "version_schema": CSV_SCHEMA_VERSION,
                    "bank_code": bank,
                    "quarter_from": q_from,
                    "quarter_to": q_to,
                    "year": year_val,
                    "change_id": _sanitize_cell(base.get("change_id", "")),
                    "section": _sanitize_cell(base.get("section", "")),
                    "table_name": _sanitize_cell(base.get("table_name", "")),
                    "table_number": _sanitize_cell(base.get("table_number", "")),
                    "table_id_t1": _sanitize_cell(base.get("table_id_t1", "")),
                    "table_id_t2": _sanitize_cell(base.get("table_id_t2", "")),
                    "page_t1": _format_cell(base.get("page_t1")),
                    "page_t2": _format_cell(base.get("page_t2")),
                    "table_status": _sanitize_cell(base.get("table_status", "")),
                    "indicator_name": _sanitize_cell(ind_name),
                    "indicator_type": _sanitize_cell(ind_type),
                    "indicator_from": _sanitize_cell(from_val),
                    "indicator_to": _sanitize_cell(to_val),
                    "resume_court": _build_resume_court(
                        ind_type,
                        ind_name,
                        from_val,
                        to_val,
                        base.get("table_status", ""),
                    ),
                    "confidence": _format_cell(base.get("confidence", "")),
                    "match_method": _sanitize_cell(base.get("match_method", "")),
                    "review_status": _sanitize_cell(
                        ind.get("review_status", base.get("review_status", ""))
                    ),
                    "idee": "",
                    "sujet": "",
                    "validation_finale": "",
                    "comment": _sanitize_cell(base.get("comment", "")),
                }
                writer.writerow(row)

    return CSV_BOM + buffer.getvalue()


def export_review_items_json_fr(
    items: list[ReviewItem],
    *,
    metadata: dict[str, Any] | None = None,
) -> str:
    """Serialize review items to JSON (French-oriented schema)."""
    payload = {
        "metadata": metadata or {},
        "total": len(items),
        "items": [item.to_dict() for item in items],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)
